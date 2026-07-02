"""
기존 Excel 보고서에서 데이터를 추출하는 모듈.
프로토타입에서는 Naver API 대신 이 모듈로 샘플 데이터를 읽는다.
"""
import re
from datetime import date
from pathlib import Path

import openpyxl

from collectors.base import (
    ClientReport,
    DailyStats,
    KeywordStats,
    MediaStats,
    MonthlyStats,
)


def _safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _safe_int(val, default=0) -> int:
    if val is None:
        return default
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _parse_cover(ws) -> dict:
    """표지 시트에서 클라이언트 메타정보 추출"""
    info = {"title": "", "homepage": "", "manager": "", "email": "", "phone": ""}
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if cell is None:
                continue
            label = str(cell).strip()
            if "홈페이지" in label:
                info["homepage"] = str(row[i + 1] or "").strip()
            elif "담 당 자" in label or "담당자" in label:
                info["manager"] = str(row[i + 1] or "").strip()
            elif "이 메 일" in label or "이메일" in label:
                info["email"] = str(row[i + 1] or "").strip()
            elif "연 락 처" in label or "연락처" in label:
                info["phone"] = str(row[i + 1] or "").strip()
    # 제목 행 (row 11)
    for row in ws.iter_rows(min_row=10, max_row=12, values_only=True):
        if row[0] and str(row[0]).strip() and str(row[0]).strip() != "Marketing Report":
            info["title"] = str(row[0]).strip()
            break
    return info


def _find_header_row(ws, keywords: list[str]) -> int | None:
    """지정 키워드가 모두 포함된 행 번호 반환"""
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip() if c else "" for c in row]
        if all(any(kw in c for c in cells) for kw in keywords):
            return i
    return None


def _row_to_monthly(row_vals: list, headers: list[str]) -> MonthlyStats | None:
    """헤더와 행 값으로 MonthlyStats 생성"""
    h = {h: i for i, h in enumerate(headers)}

    def col(name_fragments: list[str], default=None):
        for frag in name_fragments:
            for key, idx in h.items():
                if frag in key:
                    v = row_vals[idx] if idx < len(row_vals) else None
                    if v is not None:
                        return v
        return default

    month_label = col(["월별", "월"])
    if not month_label or str(month_label).strip() in ("월별", "합계", ""):
        return None

    label = str(month_label).strip()
    year, month = _parse_month_label(label)
    if month is None:
        return None
    if year is None:
        year = 0  # _infer_years에서 채워줌

    impressions = _safe_int(col(["노출수"]))
    clicks = _safe_int(col(["클릭수"]))
    ctr = _safe_float(col(["클릭률"]))
    cpc = _safe_float(col(["클릭당비용"]))
    spend = _safe_int(col(["광고비용", "광고비"]))
    avg_rank = _safe_float(col(["평균순위"]))
    conversions = _safe_int(col(["전환수"]))
    conversion_rate = _safe_float(col(["전환율"]))
    revenue = _safe_int(col(["전환매출"]))
    cpa = _safe_float(col(["전환당비용"]))
    roas = _safe_float(col(["광고수익률"]))

    return MonthlyStats(
        year=year,
        month=month,
        impressions=impressions,
        clicks=clicks,
        ctr=ctr,
        cpc=cpc,
        spend=spend,
        avg_rank=avg_rank,
        conversions=conversions,
        conversion_rate=conversion_rate,
        revenue=revenue,
        cpa=cpa,
        roas=roas,
    )


def _parse_month_label(label: str) -> tuple[int | None, int | None]:
    """'26년5월', '5월', '2026-05' 등 → (year, month)"""
    label = label.strip()
    # "26년5월" or "26년 5월"
    m = re.match(r"(\d{2,4})년\s*(\d{1,2})월", label)
    if m:
        y = int(m.group(1))
        mo = int(m.group(2))
        if y < 100:
            y += 2000
        return y, mo
    # "5월" (단순 월만 있는 경우)
    m = re.match(r"(\d{1,2})월$", label)
    if m:
        return None, int(m.group(1))
    # "2026-05"
    m = re.match(r"(\d{4})-(\d{2})", label)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def _parse_monthly_sheet(ws, report_year: int = 2026, report_month: int = 5) -> tuple[list[MonthlyStats], str]:
    """월별 시트에서 히스토리 데이터와 최신 코멘트 추출"""
    header_row_idx = _find_header_row(ws, ["월별", "노출수", "클릭수"])
    if header_row_idx is None:
        return [], ""

    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else "" for c in rows[header_row_idx - 1]]

    raw_stats = []
    comment = ""
    for row in rows[header_row_idx:]:
        if not any(c is not None for c in row):
            continue
        for cell in row:
            if cell and isinstance(cell, str) and len(cell) > 50 and "\n" in cell:
                comment = cell.strip()
                break
        stat = _row_to_monthly(list(row), headers)
        if stat:
            raw_stats.append(stat)

    # 연도 없는 항목은 보고 월 기준으로 역순 추론
    history = _infer_years(raw_stats, report_year, report_month)
    return history, comment


def _infer_years(stats: list[MonthlyStats], report_year: int, report_month: int) -> list[MonthlyStats]:
    """year=0 인 항목에 연도를 순서 기반으로 채워 넣기"""
    if not stats:
        return stats
    # 이미 연도가 있는 항목이 있으면 그걸 기준으로 역산
    # 없으면 report_year 기준으로 순서대로 채움
    result = []
    # 마지막 항목부터 역산
    current_year = report_year
    current_month = report_month
    for stat in reversed(stats):
        if stat.year and stat.year > 0:
            current_year = stat.year
            current_month = stat.month
            result.append(stat)
        else:
            # year가 0이면 current_month 기준으로 추론
            # 같은 month라고 가정하고 current_year/month 앞에 배치
            inferred_month = stat.month if stat.month else current_month
            if inferred_month <= current_month:
                inferred_year = current_year
            else:
                inferred_year = current_year - 1
            result.append(MonthlyStats(
                year=inferred_year,
                month=inferred_month,
                impressions=stat.impressions,
                clicks=stat.clicks,
                ctr=stat.ctr,
                cpc=stat.cpc,
                spend=stat.spend,
                avg_rank=stat.avg_rank,
                conversions=stat.conversions,
                conversion_rate=stat.conversion_rate,
                revenue=stat.revenue,
                cpa=stat.cpa,
                roas=stat.roas,
            ))
            current_year = inferred_year
            current_month = inferred_month - 1
            if current_month == 0:
                current_month = 12
                current_year -= 1
    result.reverse()
    return result


def _parse_media_section(ws) -> list[MediaStats]:
    """매체별 시트에서 매체별 실적 추출"""
    results = []
    header_row_idx = _find_header_row(ws, ["노출수", "클릭수", "광고비용"])
    if header_row_idx is None:
        return results

    rows = list(ws.iter_rows(values_only=True))
    section_headers = [str(c).strip() if c else "" for c in rows[header_row_idx - 1]]

    for row in rows[header_row_idx:]:
        if not any(c is not None for c in row):
            continue
        label = str(row[1] or "").strip()
        if not label or label in ("합계", "매체", "상품", "캠페인"):
            continue
        if label.startswith("■") or label.startswith("*") or label.startswith("디바이스"):
            continue
        # 숫자 데이터가 있는 행만
        if not any(isinstance(c, (int, float)) for c in row[2:8]):
            continue

        h = section_headers
        idx = {s: i for i, s in enumerate(h)}

        def col(frags, default=0):
            for f in frags:
                for k, i in idx.items():
                    if f in k and i < len(row):
                        v = row[i]
                        if v is not None:
                            return v
            return default

        results.append(
            MediaStats(
                media_name=label,
                impressions=_safe_int(col(["노출수"])),
                clicks=_safe_int(col(["클릭수"])),
                ctr=_safe_float(col(["클릭률"])),
                cpc=_safe_float(col(["클릭당비용"])),
                spend=_safe_int(col(["광고비용", "광고비"])),
                avg_rank=_safe_float(col(["평균순위"])),
                conversions=_safe_int(col(["전환수"])),
                conversion_rate=_safe_float(col(["전환율"])),
                revenue=_safe_int(col(["전환매출"])),
                cpa=_safe_float(col(["전환당비용"])),
                roas=_safe_float(col(["광고수익률"])),
            )
        )
    return results


def _parse_daily_sheet(ws, media_name: str) -> list[DailyStats]:
    """일별 시트에서 DailyStats 목록 추출"""
    header_row_idx = _find_header_row(ws, ["일별", "노출수"])
    if header_row_idx is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else "" for c in rows[header_row_idx - 1]]
    h = {s: i for i, s in enumerate(headers)}

    def col(row, frags, default=0):
        for f in frags:
            for k, i in h.items():
                if f in k and i < len(row):
                    v = row[i]
                    if v is not None:
                        return v
        return default

    daily = []
    for row in rows[header_row_idx:]:
        date_val = row[h.get("일별", 1)] if "일별" in h else row[1]
        if date_val is None or str(date_val).strip() in ("합계", "일별", ""):
            continue
        if hasattr(date_val, "date"):
            d = date_val.date()
        else:
            continue

        weekday = str(col(row, ["요일"], "")).strip()
        daily.append(
            DailyStats(
                date=d,
                weekday=weekday,
                impressions=_safe_int(col(row, ["노출수"])),
                clicks=_safe_int(col(row, ["클릭수"])),
                ctr=_safe_float(col(row, ["클릭률"])),
                cpc=_safe_float(col(row, ["클릭당비용"])),
                spend=_safe_int(col(row, ["광고비용", "광고비"])),
                avg_rank=_safe_float(col(row, ["평균순위"])),
                conversions=_safe_int(col(row, ["전환수"])),
                conversion_rate=_safe_float(col(row, ["전환율"])),
                revenue=_safe_int(col(row, ["전환매출"])),
                cpa=_safe_float(col(row, ["전환당비용"])),
                roas=_safe_float(col(row, ["광고수익률"])),
            )
        )
    return daily


def _parse_keyword_sheet(ws) -> list[KeywordStats]:
    """키워드별 시트에서 상위 50개 추출"""
    header_row_idx = _find_header_row(ws, ["키워드", "노출수"])
    if header_row_idx is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else "" for c in rows[header_row_idx - 1]]
    h = {s: i for i, s in enumerate(headers)}

    def col(row, frags, default=0):
        for f in frags:
            for k, i in h.items():
                if f in k and i < len(row):
                    v = row[i]
                    if v is not None:
                        return v
        return default

    keywords = []
    for row in rows[header_row_idx : header_row_idx + 50]:
        kw_val = row[h.get("키워드", 1)] if "키워드" in h else None
        if kw_val is None or str(kw_val).strip() in ("합계", "키워드", ""):
            continue
        keywords.append(
            KeywordStats(
                keyword=str(kw_val).strip(),
                impressions=_safe_int(col(row, ["노출수"])),
                clicks=_safe_int(col(row, ["클릭수"])),
                ctr=_safe_float(col(row, ["클릭률"])),
                cpc=_safe_float(col(row, ["클릭당비용"])),
                spend=_safe_int(col(row, ["광고비용", "광고비"])),
                avg_rank=_safe_float(col(row, ["평균순위"])),
                conversions=_safe_int(col(row, ["전환수"])),
                conversion_rate=_safe_float(col(row, ["전환율"])),
                revenue=_safe_int(col(row, ["전환매출"])),
                roas=_safe_float(col(row, ["광고수익률"])),
            )
        )
    return keywords


def read_report(filepath: str | Path) -> ClientReport:
    """Excel 보고서 파일 전체를 파싱해서 ClientReport로 반환"""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    sheet_names = wb.sheetnames

    # 표지
    cover = _parse_cover(wb["표지"]) if "표지" in sheet_names else {}

    # 파일명에서 연도/월 추출 (연도 추론용)
    fname_stem = Path(filepath).stem
    year_hint, month_hint = 2026, 5
    m = re.search(r"(\d{4})년.{0,2}(\d{1,2})월", fname_stem)
    if m:
        year_hint, month_hint = int(m.group(1)), int(m.group(2))

    # 월별
    history, comment = [], ""
    for sname in ("월별", "토탈(브검포함)"):
        if sname in sheet_names:
            history, comment = _parse_monthly_sheet(wb[sname], year_hint, month_hint)
            break

    # 최신 월 (history 마지막)
    monthly_total = history[-1] if history else None

    # 매체별
    media_list = []
    if "매체별" in sheet_names:
        media_list = _parse_media_section(wb["매체별"])

    # 일별/키워드별 데이터를 매체 이름으로 매핑
    media_map = {m.media_name: m for m in media_list}

    daily_sheet_map = {
        "네이버 쇼핑검색": ["N_쇼핑_일별", "N_쇼핑검색_일별"],
        "네이버 파워링크": ["N_파워링크_일별"],
        "GFA 카탈로그": ["N_카탈로그_일별"],
        "GFA 애드부스트": ["N_애드부스트_일별", "N_GFA_애드부스트 일별"],
        "GFA": ["N_GFA_일별"],
    }
    keyword_sheet_map = {
        "네이버 쇼핑검색": ["N_쇼핑_키워드별MO", "N_쇼핑검색_검색어별"],
        "네이버 파워링크": ["N_파워링크_키워드별"],
    }

    for media_name, media_stat in media_map.items():
        for mapped_name, sheet_candidates in daily_sheet_map.items():
            if mapped_name in media_name or media_name in mapped_name:
                for sname in sheet_candidates:
                    if sname in sheet_names:
                        media_stat.daily_stats = _parse_daily_sheet(wb[sname], media_name)
                        break
        for mapped_name, sheet_candidates in keyword_sheet_map.items():
            if mapped_name in media_name or media_name in mapped_name:
                for sname in sheet_candidates:
                    if sname in sheet_names:
                        media_stat.keyword_stats = _parse_keyword_sheet(wb[sname])
                        break

    # 파일명에서 클라이언트명 추출 (fallback)
    fname = Path(filepath).stem
    client_name = cover.get("title", fname)
    # "위폭스 5월 운영데이터 보고서" → "위폭스" 추출
    client_name = client_name.split()[0] if client_name else fname

    year = monthly_total.year if monthly_total else 2026
    month = monthly_total.month if monthly_total else 5

    return ClientReport(
        client_name=client_name,
        homepage=cover.get("homepage", ""),
        manager=cover.get("manager", ""),
        email=cover.get("email", ""),
        phone=cover.get("phone", ""),
        year=year,
        month=month,
        monthly_total=monthly_total,
        monthly_history=history,
        media_breakdown=media_list,
        comment=comment,
    )
