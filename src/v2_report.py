"""
버전B 보고서 생성기 — 미래아이엔씨 자동보고서 양식 기반
RAW 시트에 네이버 API 데이터를 채우면 엑셀 수식이 자동으로 Summary/상세 시트를 계산.

흐름:
  1. 네이버 API에서 일별 × 기기 × 캠페인 RAW 데이터 수집
  2. 키워드 데이터 수집
  3. 템플릿 복사 → 숨김 시트 모두 표시
  4. 파워링크 RAW 시트에 데이터 기입
  5. Summary/키워드 시트 날짜·기간 갱신
  6. 엑셀 수식이 나머지 자동 계산
"""
import calendar
import json
import shutil
import time
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

from collectors.naver_searchad import (
    _get, _batch_stats, _row_metrics, _collect_keywords,
    MOCK_MODE, _ACCT_BY_ID, _STAT_FIELDS,
)

_ROOT = Path(__file__).resolve().parents[1]
V2_TEMPLATE = _ROOT / "미래아이엔씨 자동보고서 양식_페이파란_Weekly Report_260310.xlsx"
OUTPUT_DIR = _ROOT / "output"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  1. 데이터 수집
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _device_from_name(name: str) -> str:
    """캠페인/광고그룹 이름에서 디바이스 추출.
    _PC → PC, _MO → 모바일, 그 외 → 전체.

    참고: /stats 의 breakdown=pcMblTp 는 '최근 7일'만 조회 가능(에러 11004)해
    지난 달 전체를 다루는 월간 보고서에는 쓸 수 없다. 네이버 검색광고는
    캠페인/광고그룹을 PC/MO 로 분리 운영하는 것이 일반적이라(_PC/_MO 접미사)
    이름 기반 분류가 월간 단위에서는 정확하고 안정적이다."""
    upper = name.upper()
    if "_PC" in upper:
        return "PC"
    if "_MO" in upper:
        return "모바일"
    return "전체"


def _collect_raw_powerlink(customer_id: str, year: int, month: int) -> tuple:
    """파워링크 RAW 데이터: 일별 × 캠페인 단위 수집.
    캠페인명(_PC/_MO)에서 디바이스 분류.
    Returns: (raw_rows, camp_ids, camp_name_map)
    raw_rows: [[date_str, device, campaign, "파워링크", imp, clk, spend, conv, revenue], ...]
    """
    campaigns = _get("/ncc/campaigns", customer_id=customer_id)
    # 파워링크(WEB_SITE)만 — 브랜드검색(BRAND_SEARCH)·쇼핑(SHOPPING) 등 타 매체 제외
    campaigns = [c for c in campaigns if c.get("campaignTp") == "WEB_SITE"]
    camp_ids = [c["nccCampaignId"] for c in campaigns]
    camp_name = {c["nccCampaignId"]: c.get("name", "") for c in campaigns}

    if not camp_ids:
        return [], [], {}

    last_day = calendar.monthrange(year, month)[1]
    FIELDS = json.dumps(_STAT_FIELDS)

    def fetch_batch(batch, ds, date_str):
        """단일 배치 조회 → 행 리스트. 일시적 오류는 재시도, 최종 실패 시 예외 전파."""
        path = (
            "/stats?ids=" + urllib.parse.quote(",".join(batch))
            + "&fields=" + urllib.parse.quote(FIELDS)
            + "&timeRange=" + urllib.parse.quote(json.dumps({"since": ds, "until": ds}))
        )
        last_err = None
        for attempt in range(4):
            try:
                r = _get(path, customer_id=customer_id)
                data = r.get("data", []) if isinstance(r, dict) else r
                rows = []
                for row in data:
                    cid = row.get("id", "")
                    cname = camp_name.get(cid, cid)
                    device = _device_from_name(cname)
                    imp = int(row.get("impCnt", 0) or 0)
                    clk = int(row.get("clkCnt", 0) or 0)
                    spend = int(row.get("salesAmt", 0) or 0)
                    conv = int(row.get("ccnt", 0) or 0)
                    rev = int(row.get("convAmt", 0) or 0)
                    if imp or clk or spend:
                        rows.append(
                            [date_str, device, cname, "파워링크", imp, clk, spend, conv, rev]
                        )
                return rows
            except Exception as e:
                last_err = e
                time.sleep(0.6 * (attempt + 1))  # backoff (rate limit 완화)
        raise RuntimeError(f"stats 조회 실패({ds}): {last_err}")

    def fetch_day(day):
        ds = f"{year}-{month:02d}-{day:02d}"
        date_str = f"{year}.{month:02d}.{day:02d}."
        day_rows = []
        BATCH = 10
        for i in range(0, len(camp_ids), BATCH):
            batch = camp_ids[i : i + BATCH]
            day_rows.extend(fetch_batch(batch, ds, date_str))
        return day_rows

    days = list(range(1, last_day + 1))
    all_rows = []
    failed_days = []

    # 1차: 병렬 수집(동시성 4로 rate limit 완화). 실패한 날은 모아둔다.
    with ThreadPoolExecutor(max_workers=4) as ex:
        futures = {ex.submit(fetch_day, d): d for d in days}
        for fut in as_completed(futures):
            d = futures[fut]
            try:
                all_rows.extend(fut.result())
            except Exception:
                failed_days.append(d)

    # 2차: 실패한 날만 순차 재시도(부하 없이). 그래도 실패하면 어느 날인지 명시하고 중단.
    for d in sorted(failed_days):
        try:
            all_rows.extend(fetch_day(d))
        except Exception as e:
            raise RuntimeError(f"{year}-{month:02d}-{d:02d} 수집 실패(복구 불가): {e}")

    return all_rows, camp_ids, camp_name


def _collect_raw_creative(
    customer_id: str, camp_ids: list, camp_name: dict, time_range: dict,
    top_n: int = 150,
) -> list:
    """소재(광고)별 성과를 캠페인/광고그룹 계층과 함께 수집 (기간 집계).
    Returns: [{ad_id, campaign, adgroup, name, impressions, clicks, spend, conversions}, ...]
    """
    MAX_ADS = 1000
    ad_meta: dict[str, dict] = {}  # ad_id → {campaign, adgroup, name}

    for cid in camp_ids:
        if len(ad_meta) >= MAX_ADS:
            break
        cname = camp_name.get(cid, cid)
        try:
            adgroups = _get(f"/ncc/adgroups?nccCampaignId={cid}", customer_id=customer_id)
        except Exception:
            continue
        for ag in adgroups:
            if len(ad_meta) >= MAX_ADS:
                break
            agid = ag.get("nccAdgroupId")
            agname = ag.get("name", "")
            if not agid:
                continue
            try:
                ads = _get(f"/ncc/ads?nccAdgroupId={agid}", customer_id=customer_id)
            except Exception:
                continue
            for ad in ads:
                aid = ad.get("nccAdId")
                if not aid:
                    continue
                ad_meta[aid] = {
                    "campaign": cname,
                    "adgroup": agname,
                    "name": _ad_display_name(ad),
                }

    if not ad_meta:
        return []

    rows = _batch_stats(list(ad_meta.keys()), time_range, customer_id=customer_id)
    result = []
    for d in rows:
        imp = int(d.get("impCnt", 0) or 0)
        clk = int(d.get("clkCnt", 0) or 0)
        if imp == 0 and clk == 0:
            continue
        meta = ad_meta.get(d.get("id"), {})
        result.append({
            "ad_id": d.get("id", ""),
            "campaign": meta.get("campaign", ""),
            "adgroup": meta.get("adgroup", ""),
            "name": meta.get("name", d.get("id", "")),
            "impressions": imp,
            "clicks": clk,
            "spend": int(d.get("salesAmt", 0) or 0),
            "conversions": int(d.get("ccnt", 0) or 0),
        })
    result.sort(key=lambda x: -x["spend"])
    return result[:top_n]


def _ad_display_name(ad: dict) -> str:
    """광고(소재) 객체에서 표시용 소재명 추출.
    파워링크 소재는 제목(headline)을 우선 사용, 없으면 소재 ID."""
    detail = ad.get("ad", {}) or {}
    for section in ("pc", "mobile"):
        blk = detail.get(section) or {}
        for k in ("headline", "title", "subject"):
            if blk.get(k):
                return str(blk[k])
    for k in ("headline", "title", "subject"):
        if detail.get(k):
            return str(detail[k])
    return ad.get("nccAdId", "")


def _collect_keywords_with_hierarchy(
    customer_id: str, camp_ids: list, camp_name: dict,
    time_range: dict, top_n: int = 50,
) -> list:
    """키워드별 성과를 캠페인/광고그룹 계층 정보와 함께 수집.
    Returns: [{keyword, campaign, adgroup, impressions, clicks, ...}, ...]
    """
    MAX_KEYWORDS = 800
    kw_meta: dict[str, dict] = {}  # keyword_id → {keyword, campaign, adgroup}

    for cid in camp_ids:
        if len(kw_meta) >= MAX_KEYWORDS:
            break
        cname = camp_name.get(cid, cid)
        try:
            adgroups = _get(f"/ncc/adgroups?nccCampaignId={cid}", customer_id=customer_id)
        except Exception:
            continue
        for ag in adgroups:
            if len(kw_meta) >= MAX_KEYWORDS:
                break
            agid = ag.get("nccAdgroupId")
            agname = ag.get("name", "")
            if not agid:
                continue
            try:
                kws = _get(f"/ncc/keywords?nccAdgroupId={agid}", customer_id=customer_id)
            except Exception:
                continue
            for kw in kws:
                kid = kw.get("nccKeywordId")
                if kid:
                    kw_meta[kid] = {
                        "keyword": kw.get("keyword", kid),
                        "campaign": cname,
                        "adgroup": agname,
                    }

    if not kw_meta:
        return []

    rows = _batch_stats(list(kw_meta.keys()), time_range, customer_id=customer_id)
    result = []
    for d in rows:
        m = _row_metrics(d)
        if m["impressions"] == 0 and m["clicks"] == 0:
            continue
        meta = kw_meta.get(d.get("id"), {})
        m["keyword"] = meta.get("keyword", d.get("id", ""))
        m["campaign"] = meta.get("campaign", "")
        m["adgroup"] = meta.get("adgroup", "")
        result.append(m)

    result.sort(key=lambda x: -x["impressions"])
    return result[:top_n]


def _month_day_spans(year: int, month: int) -> tuple:
    """월을 요일 기준 연속 구간으로 분할.
    Returns (weekday_spans, weekend_spans) — 각 span=(시작일, 끝일) 연속 구간.
    평일(월~금)/주말(토·일)이 각각 연속 구간이므로 _batch_stats를 구간 단위로 조회."""
    last = calendar.monthrange(year, month)[1]
    weekday_spans, weekend_spans = [], []
    d = 1
    while d <= last:
        is_wknd = calendar.weekday(year, month, d) >= 5
        start = d
        while d <= last and (calendar.weekday(year, month, d) >= 5) == is_wknd:
            d += 1
        (weekend_spans if is_wknd else weekday_spans).append((start, d - 1))
    return weekday_spans, weekend_spans


def _collect_keywords_split(
    customer_id: str, camp_ids: list, camp_name: dict,
    year: int, month: int, top_n: int = 50,
) -> tuple:
    """키워드 성과를 평일/주말로 분리 수집.
    camp_ids는 '광고비 내림차순'으로 정렬돼 들어온다고 가정(상위 캠페인 우선).
    흐름: (1) 캠페인당 상한을 둬 키워드 메타를 폭넓게 수집,
          (2) 월간 성과로 상위 top_n 키워드 선별,
          (3) 그 상위 키워드만 평일/주말 구간으로 분리 집계(효율).
    Returns: (weekday_keywords, weekend_keywords)
    """
    PER_CAMP = 200
    kw_meta: dict[str, dict] = {}
    for cid in camp_ids:
        cname = camp_name.get(cid, cid)
        try:
            adgroups = _get(f"/ncc/adgroups?nccCampaignId={cid}", customer_id=customer_id)
        except Exception:
            continue
        cnt = 0
        for ag in adgroups:
            if cnt >= PER_CAMP:
                break
            agid = ag.get("nccAdgroupId")
            if not agid:
                continue
            agname = ag.get("name", "")
            try:
                kws = _get(f"/ncc/keywords?nccAdgroupId={agid}", customer_id=customer_id)
            except Exception:
                continue
            for kw in kws:
                kid = kw.get("nccKeywordId")
                if not kid:
                    continue
                kw_meta[kid] = {
                    "keyword": kw.get("keyword", kid),
                    "campaign": cname,
                    "adgroup": agname,
                }
                cnt += 1
                if cnt >= PER_CAMP:
                    break

    if not kw_meta:
        return [], []

    # ── 월간 성과로 상위 top_n 키워드 선별 (광고비 기준) ──
    last_day = calendar.monthrange(year, month)[1]
    tr_month = {"since": f"{year}-{month:02d}-01", "until": f"{year}-{month:02d}-{last_day:02d}"}
    try:
        month_rows = _batch_stats(list(kw_meta.keys()), tr_month, customer_id=customer_id)
    except Exception:
        month_rows = []
    spend_by = {}
    for d in month_rows:
        kid = d.get("id")
        if kid:
            spend_by[kid] = int(d.get("salesAmt", 0) or 0)
    top_ids = [kid for kid, _ in sorted(spend_by.items(), key=lambda x: -x[1])[:top_n]]
    if not top_ids:
        return [], []
    top_set = set(top_ids)

    def aggregate_over(spans: list) -> list:
        acc: dict[str, dict] = {}
        for (s, e) in spans:
            tr = {"since": f"{year}-{month:02d}-{s:02d}", "until": f"{year}-{month:02d}-{e:02d}"}
            try:
                rows = _batch_stats(top_ids, tr, customer_id=customer_id)
            except Exception:
                continue
            for d in rows:
                kid = d.get("id")
                if kid not in top_set:
                    continue
                m = _row_metrics(d)
                if m["impressions"] == 0 and m["clicks"] == 0:
                    continue
                a = acc.setdefault(kid, {
                    "impressions": 0, "clicks": 0, "spend": 0, "conversions": 0,
                    "_rank_w": 0.0, "_rank_imp": 0,
                })
                a["impressions"] += m["impressions"]
                a["clicks"] += m["clicks"]
                a["spend"] += m["spend"]
                a["conversions"] += m["conversions"]
                a["_rank_w"] += (m.get("avg_rank", 0) or 0) * m["impressions"]
                a["_rank_imp"] += m["impressions"]
        result = []
        for kid, a in acc.items():
            meta = kw_meta.get(kid, {})
            result.append({
                "keyword": meta.get("keyword", kid),
                "campaign": meta.get("campaign", ""),
                "adgroup": meta.get("adgroup", ""),
                "impressions": a["impressions"],
                "clicks": a["clicks"],
                "spend": a["spend"],
                "conversions": a["conversions"],
                "avg_rank": round(a["_rank_w"] / a["_rank_imp"], 1) if a["_rank_imp"] else 0,
            })
        result.sort(key=lambda x: -x["spend"])
        return result

    weekday_spans, weekend_spans = _month_day_spans(year, month)
    return aggregate_over(weekday_spans), aggregate_over(weekend_spans)


def _split_period_totals(raw_rows: list, year: int, month: int) -> tuple:
    """RAW를 요일 기준으로 평일/주말 캠페인 총합으로 집계.
    키워드 시트의 '합계'를 월(기간) 총합과 정합시키기 위한 기준값.
    Returns: (weekday_total, weekend_total) — 각 {imp, clk, spend, conv}
    """
    wd = {"imp": 0, "clk": 0, "spend": 0, "conv": 0}
    we = {"imp": 0, "clk": 0, "spend": 0, "conv": 0}
    for row in raw_rows:
        date_str, _, _, _, imp, clk, spend, conv = row[:8]
        try:
            day = int(date_str.replace(".", " ").split()[2])
        except (IndexError, ValueError):
            continue
        t = we if calendar.weekday(year, month, day) >= 5 else wd
        t["imp"] += imp; t["clk"] += clk; t["spend"] += spend; t["conv"] += conv
    return wd, we


def _merge_keywords(*lists) -> list:
    """여러 키워드 리스트를 (campaign, adgroup, keyword) 기준으로 합산해 월간 집계 생성.
    코멘트용 — impressions 내림차순."""
    acc: dict[tuple, dict] = {}
    for lst in lists:
        for kw in lst:
            key = (kw.get("campaign", ""), kw.get("adgroup", ""), kw.get("keyword", ""))
            a = acc.setdefault(key, {
                "keyword": kw.get("keyword", ""), "campaign": kw.get("campaign", ""),
                "adgroup": kw.get("adgroup", ""),
                "impressions": 0, "clicks": 0, "spend": 0, "conversions": 0,
            })
            a["impressions"] += kw.get("impressions", 0) or 0
            a["clicks"] += kw.get("clicks", 0) or 0
            a["spend"] += kw.get("spend", 0) or 0
            a["conversions"] += kw.get("conversions", 0) or 0
    result = list(acc.values())
    result.sort(key=lambda x: -x["impressions"])
    return result


def _mock_raw_powerlink(year: int, month: int) -> list:
    """Mock RAW 데이터 생성"""
    import random

    random.seed(year * 100 + month)
    last_day = calendar.monthrange(year, month)[1]
    campaigns = ["캠페인_브랜드", "캠페인_일반키워드", "캠페인_상품명"]
    devices = ["PC", "모바일"]
    rows = []
    for day in range(1, last_day + 1):
        date_str = f"{year}.{month:02d}.{day:02d}."
        for camp in campaigns:
            for device in devices:
                mo = 1.5 if device == "모바일" else 1.0
                imp = int(random.randint(500, 3000) * mo)
                clk = int(random.randint(10, 80) * mo)
                spend = int(random.randint(5000, 50000) * mo)
                conv = random.randint(0, 5)
                rev = conv * random.randint(20000, 80000)
                rows.append([date_str, device, camp, "파워링크", imp, clk, spend, conv, rev])
    return rows


def _mock_creative(year: int, month: int) -> list:
    """Mock 소재별 데이터"""
    import random

    random.seed(year * 100 + month + 13)
    campaigns = ["캠페인_브랜드_MO", "캠페인_일반키워드_MO", "캠페인_브랜드_PC"]
    adgroups = ["a.메인_브랜드", "b.세부_브랜드", "c.서브_일반"]
    result = []
    for i in range(24):
        imp = random.randint(100, 5000)
        clk = random.randint(3, 120)
        sp = random.randint(10000, 200000)
        cv = random.randint(0, 8)
        result.append({
            "ad_id": f"nad-a001-01-{260000000000 + i:012d}",
            "campaign": campaigns[i % len(campaigns)],
            "adgroup": adgroups[i % len(adgroups)],
            "name": f"소재 {i + 1}",
            "impressions": imp, "clicks": clk, "spend": sp, "conversions": cv,
        })
    result.sort(key=lambda x: -x["spend"])
    return result


def _mock_keyword_data(year: int, month: int) -> list:
    """Mock 키워드 데이터 (캠페인/광고그룹 포함)"""
    import random

    random.seed(year * 100 + month + 7)
    campaigns = ["캠페인_브랜드_PC", "캠페인_브랜드_MO", "캠페인_일반키워드_PC"]
    adgroups = ["a.메인_브랜드", "b.세부_브랜드", "c.서브_일반"]
    kw_list = [
        "브랜드명", "브랜드명 추천", "강아지 사료", "반려견 간식", "유기농 사료",
        "수제 간식", "대형견 사료", "퍼피 사료", "관절 영양제", "치석 제거",
        "강아지 영양제", "노견 사료", "습식 사료", "사료 추천", "강아지 용품",
        "반려동물", "자연사료", "생식사료", "다이어트 사료", "알러지 사료",
    ]
    result = []
    for i, kw in enumerate(kw_list):
        imp = random.randint(1000, 30000)
        clk = random.randint(20, 500)
        sp = random.randint(10000, 200000)
        cv = random.randint(0, 15)
        result.append({
            "keyword": kw,
            "campaign": campaigns[i % len(campaigns)],
            "adgroup": adgroups[i % len(adgroups)],
            "impressions": imp, "clicks": clk,
            "ctr": round(clk / imp, 4) if imp else 0,
            "cpc": round(sp / clk) if clk else 0,
            "spend": sp,
            "avg_rank": round(random.uniform(1.0, 5.0), 1),
            "conversions": cv,
            "conversion_rate": round(cv / clk, 4) if clk else 0,
            "revenue": cv * random.randint(20000, 80000),
            "cpa": round(sp / cv) if cv else 0,
            "roas": 0,
        })
    result.sort(key=lambda x: -x["impressions"])
    return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  2. 엑셀 템플릿 채우기
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _clear_data_rows(ws, start_row: int, max_col: int = None):
    """start_row 부터 기존 데이터를 모두 지운다 (셀 서식은 유지)."""
    if max_col is None:
        max_col = ws.max_column
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _fill_raw_sheet(wb, sheet_name: str, rows: list):
    """파워링크 RAW 시트의 기존 데이터를 지우고 새 데이터로 채움.
    RAW 칼럼: A일자 B기기 C캠페인 D매체 E노출 F클릭 G광고비 H구매수 I매출액
    rows: [[date_str, device, campaign, media, imp, clk, spend, conv, revenue], ...]
    (revenue 없는 8칼럼 행도 허용)
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    _clear_data_rows(ws, 2, 9)
    for i, row_data in enumerate(rows):
        r = i + 2
        for j, val in enumerate(row_data):
            ws.cell(row=r, column=j + 1, value=val)


def _fill_creative_raw(wb, sheet_name: str, creatives: list):
    """파링_소재 RAW 시트 채우기.
    칼럼: A소재ID B캠페인 C광고그룹 D노출수 E클릭수 F총비용 G전환수 I소재ID(누적) J소재명
    I/J 는 소재별_주간 시트의 소재명 VLOOKUP 참조 테이블로 쓰임.
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    _clear_data_rows(ws, 2, 10)
    for i, cr in enumerate(creatives):
        r = i + 2
        ws.cell(row=r, column=1, value=cr.get("ad_id", ""))       # A: 소재ID
        ws.cell(row=r, column=2, value=cr.get("campaign", ""))     # B: 캠페인
        ws.cell(row=r, column=3, value=cr.get("adgroup", ""))      # C: 광고그룹
        ws.cell(row=r, column=4, value=cr.get("impressions", 0))   # D: 노출수
        ws.cell(row=r, column=5, value=cr.get("clicks", 0))        # E: 클릭수
        ws.cell(row=r, column=6, value=cr.get("spend", 0))         # F: 총비용
        ws.cell(row=r, column=7, value=cr.get("conversions", 0))   # G: 전환수
        ws.cell(row=r, column=9, value=cr.get("ad_id", ""))        # I: 소재ID(누적)
        ws.cell(row=r, column=10, value=cr.get("name", ""))        # J: 소재명


def _fill_creative_weekly(wb, sheet_name: str, creatives: list):
    """파워링크_소재별_주간 시트 채우기 (표시용 값 시트).
    데이터 시작행(합계 SUM 수식 아래)부터 소재별 값을 기입.
    칼럼: B소재 C소재명 D캠페인 E광고그룹 F노출 G클릭 H CTR I CPC J광고비 K전환수 L CVR
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    # 데이터 시작행 탐색: F열(노출) 합계 SUM 수식 바로 아래
    data_start = None
    for r in range(5, 20):
        cell = ws.cell(row=r, column=6)  # F열
        if isinstance(cell, MergedCell):
            continue
        val = cell.value
        if val and isinstance(val, str) and val.startswith("=SUM"):
            data_start = r + 1
            break
    if data_start is None:
        data_start = 11

    # 기존 데이터 지우기 (B~L 및 우측 메모 N~O)
    for r in range(data_start, min(ws.max_row + 1, data_start + 500)):
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for r in range(data_start - 1, min(ws.max_row + 1, data_start + 60)):
        for c in (14, 15):  # N, O 잔재 메모 제거
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # 작성자 의도: 값칸(B,D,E,F,G,J,K)은 값, 파생칸(C,H,I,L)은 수식.
    #  C=소재명 VLOOKUP(파링_소재 RAW I:J), H=CTR, I=CPC, L=CVR
    # 템플릿 서식행 초과분도 서식 유지 위해 첫 데이터행 서식 복사
    for i, cr in enumerate(creatives):
        r = data_start + i
        if r != data_start:
            _copy_row_style(ws, data_start, r, range(2, 13))
        ws.cell(row=r, column=2, value=cr.get("ad_id", ""))                       # B: 소재ID
        ws.cell(row=r, column=3, value=f"=VLOOKUP(B{r},'파링_소재 RAW'!I:J,2,)")   # C: 소재명(조회)
        ws.cell(row=r, column=4, value=cr.get("campaign", ""))                    # D: 캠페인
        ws.cell(row=r, column=5, value=cr.get("adgroup", ""))                     # E: 광고그룹
        ws.cell(row=r, column=6, value=cr.get("impressions", 0) or 0)             # F: 노출
        ws.cell(row=r, column=7, value=cr.get("clicks", 0) or 0)                  # G: 클릭
        ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/F{r},0)")                  # H: CTR
        ws.cell(row=r, column=9, value=f"=IFERROR(J{r}/G{r},0)")                  # I: CPC
        ws.cell(row=r, column=10, value=cr.get("spend", 0) or 0)                  # J: 광고비
        ws.cell(row=r, column=11, value=cr.get("conversions", 0) or 0)            # K: 전환수
        ws.cell(row=r, column=12, value=f"=IFERROR(K{r}/G{r},0)")                 # L: CVR


def _find_daily_start_row(ws) -> int:
    """Summary 시트에서 일별 데이터 시작 행을 찾는다.
    B열에 날짜 형식(YYYY.MM.DD) 텍스트가 처음 나타나는 행."""
    for r in range(50, min(ws.max_row + 1, 120)):
        val = ws.cell(row=r, column=2).value
        if val is None:
            continue
        s = str(val).strip()
        # "2026.03.02" 같은 패턴
        if len(s) >= 8 and s.count(".") >= 2:
            try:
                parts = s.split(".")
                if len(parts) >= 3 and parts[0].isdigit():
                    return r
            except Exception:
                pass
    return 69  # 기본 fallback


def _update_summary_dates(wb, sheet_name: str, year: int, month: int):
    """Summary 시트의 일별 영역 날짜를 새 기간으로 갱신."""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    last_day = calendar.monthrange(year, month)[1]

    daily_start = _find_daily_start_row(ws)

    # 날짜 기입 (원본 형식: "YYYY.MM.DD." — 마지막 마침표 포함)
    # 템플릿이 앞쪽 일부 행(예: 1/1~1/4)을 숨겨뒀을 수 있어, 기입하는 행은 숨김 해제
    for d in range(1, last_day + 1):
        r = daily_start + d - 1
        ws.cell(row=r, column=2, value=f"{year}.{month:02d}.{d:02d}.")
        ws.row_dimensions[r].hidden = False

    # 남은 행 비우기 (최대 40행까지)
    for r in range(daily_start + last_day, daily_start + 40):
        cell = ws.cell(row=r, column=2)
        if not isinstance(cell, MergedCell):
            cell.value = None


def _update_period_text(wb, year: int, month: int):
    """모든 시트의 기간 텍스트를 갱신:
    - '* 기간 : 2026.03.02~2026.03.06' → 해당 월 전체 기간
    - '구분\\n(1/1~1/31)' → 해당 월로 갱신
    - 주차별 레이블 (B35~B39) → 해당 월 주차로 갱신
    """
    last_day = calendar.monthrange(year, month)[1]
    new_period_full = f"{year}.{month:02d}.01~{year}.{month:02d}.{last_day:02d}"

    for ws in wb.worksheets:
        for r in range(1, 21):
            for c in range(1, 25):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if not val or not isinstance(val, str):
                    continue

                # "* 기간 : 2026.03.02~2026.03.06" 패턴만 갱신.
                # (주의) "기간"+"~" 만으로 매칭하면 '기간/~'를 포함한 코멘트(L9)까지
                # 덮어쓰므로, 반드시 '* 기간'으로 시작하는 라벨 셀만 교체한다.
                if val.lstrip().startswith("* 기간"):
                    cell.value = f"* 기간 : {new_period_full}"
                # "구분\n(1/1~1/31)" 패턴 (Summary B4)
                elif val.lstrip().startswith("구분") and "~" in val:
                    cell.value = f"구분\n({month}/1~{month}/{last_day})"

    # 주차별 블록(Total/평일/주말) 수식·레이블 전면 재생성
    _rebuild_summary_weekly(wb, "파워링크_Summary", year, month)


def _iso_weeks_of_month(last_day: int) -> list:
    """월을 일자 기준 7일 단위 주차로 분할. [(1,7),(8,14),...]"""
    weeks = []
    d = 1
    while d <= last_day:
        end = min(d + 6, last_day)
        weeks.append((d, end))
        d = end + 1
    return weeks


def _sum_expr(col: str, rows: list) -> str | int:
    """지정 행 목록에 대한 SUM 수식 문자열. 빈 목록이면 0."""
    if not rows:
        return 0
    return "=SUM(" + ",".join(f"{col}{r}" for r in rows) + ")"


def _write_weekly_row(ws, wr: int, label: str, daily_rows: list):
    """주차 블록의 한 행(wr)에 라벨 + 지표 수식을 기입.
    daily_rows: 이 주차에 해당하는 일별 데이터 행 목록(연속/비연속 모두 허용)."""
    ws.cell(row=wr, column=2, value=label)                                    # B: 라벨
    ws.cell(row=wr, column=3, value=_sum_expr("C", daily_rows))               # C: 노출
    ws.cell(row=wr, column=4, value=_sum_expr("D", daily_rows))               # D: 클릭
    ws.cell(row=wr, column=5, value=f"=IFERROR(D{wr}/C{wr},)")                # E: CTR
    ws.cell(row=wr, column=6, value=f"=IFERROR(G{wr}/D{wr},0)")               # F: CPC
    ws.cell(row=wr, column=7, value=_sum_expr("G", daily_rows))               # G: 광고비
    ws.cell(row=wr, column=8, value=_sum_expr("H", daily_rows))               # H: 전환수
    ws.cell(row=wr, column=9, value=f"=IFERROR(H{wr}/D{wr},0)")               # I: CVR
    ws.cell(row=wr, column=10, value=f"=IFERROR(G{wr}/H{wr},0)")              # J: 전환당비용


def _find_weekly_block(ws, header_kw: str, search_from: int, search_to: int):
    """주차 블록의 (데이터시작행, 전주대비행)을 찾는다.
    header_kw: '주차별 Total' / '주차별 평일' / '주차별 주말'."""
    header_row = None
    for r in range(search_from, search_to):
        val = ws.cell(row=r, column=2).value
        if isinstance(val, str) and header_kw in val:
            header_row = r
            break
    if header_row is None:
        return None, None
    data_start = header_row + 3  # 헤더 + 2줄 컬럼헤더
    delta_row = None
    for r in range(data_start, data_start + 12):
        val = ws.cell(row=r, column=2).value
        if isinstance(val, str) and "전주대비" in val:
            delta_row = r
            break
    return data_start, delta_row


def _rebuild_summary_weekly(wb, sheet_name: str, year: int, month: int):
    """Summary의 3개 주차 블록(Total/평일/주말)을 해당 월 달력 기준으로 재생성.
    - 각 주차의 SUM 범위를 실제 일별 데이터 행에 맞춰 다시 계산
    - 평일/주말 블록은 요일(월~금 / 토·일) 기준으로 일자를 분류
    - 전주대비 행은 마지막 두 주차를 참조하도록 재설정
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    last_day = calendar.monthrange(year, month)[1]
    daily_start = _find_daily_start_row(ws)
    weeks = _iso_weeks_of_month(last_day)

    def day_row(d):
        return daily_start + d - 1

    blocks = [
        ("주차별 Total", 28, 44, lambda days: days),
        ("주차별 평일", 41, 55, lambda days: [d for d in days
                                             if calendar.weekday(year, month, d) < 5]),
        ("주차별 주말", 52, 66, lambda days: [d for d in days
                                             if calendar.weekday(year, month, d) >= 5]),
    ]

    for header_kw, sf, st, day_filter in blocks:
        data_start, delta_row = _find_weekly_block(ws, header_kw, sf, st)
        if data_start is None:
            continue
        max_slots = (delta_row - data_start) if delta_row else len(weeks)

        for i in range(max_slots):
            wr = data_start + i
            if i < len(weeks):
                d1, d2 = weeks[i]
                sel = day_filter(list(range(d1, d2 + 1)))
                rows = [day_row(d) for d in sel]
                label = f"{month}/{d1} ~ {month}/{d2}"
                _write_weekly_row(ws, wr, label, rows)
                ws.row_dimensions[wr].hidden = False  # 숨겨진 첫 주 행 등 노출
            else:
                # 사용하지 않는 주차 행 비우기
                for c in range(2, 11):
                    cell = ws.cell(row=wr, column=c)
                    if not isinstance(cell, MergedCell):
                        cell.value = None

        # 전주대비 행: 마지막 두 주차 참조
        if delta_row and len(weeks) >= 2:
            cur = data_start + len(weeks) - 1
            prev = data_start + len(weeks) - 2
            ws.cell(row=delta_row, column=2, value="전주대비")
            for c in range(3, 11):
                col = openpyxl.utils.get_column_letter(c)
                ws.cell(row=delta_row, column=c,
                        value=f'=IFERROR({col}{cur}/{col}{prev}-1,"-")')


def _fill_keyword_sheet(wb, sheet_name: str, keywords: list, period_total: dict = None):
    """키워드 시트에 데이터 채우기.
    원본 템플릿 컬럼 구조 (B~N):
      C=캠페인, D=광고그룹, E=키워드, F=평균노출순위,
      G=노출, H=클릭, I=CTR, J=CPC, K=광고비,
      L=전환수, M=CVR, N=전환당비용
    period_total: 이 기간(평일/주말)의 캠페인 총합 {imp,clk,spend,conv}.
      주어지면 상위 키워드 아래 '기타(상위 외)' 행을 추가해
      합계(=상위50+기타)가 월(기간) 총합과 정합되게 한다. (전체 키워드 20만개라 상위만 표시)
    """
    if sheet_name not in wb.sheetnames or not keywords:
        return
    ws = wb[sheet_name]

    # 데이터 시작행 탐색: 집계 행(SUM 수식) 아래
    data_start = None
    for r in range(5, 30):
        # G열(노출) 또는 F열에서 SUM 수식 찾기
        for check_col in (7, 6):
            cell = ws.cell(row=r, column=check_col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if val and isinstance(val, str) and val.startswith("="):
                data_start = r + 1
                break
        if data_start:
            break
    if data_start is None:
        data_start = 10

    # 기존 데이터 지우기
    for r in range(data_start, min(ws.max_row + 1, data_start + 200)):
        for c in range(2, 17):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # 작성자 의도: 값칸(C,D,E,F,G,H,K,L)은 값, 파생칸(I,J,M,N)은 수식.
    #  I=CTR, J=CPC, M=CVR, N=전환당비용
    # 템플릿에 서식이 있는 데이터행 수보다 많이 쓰면 초과행이 무서식이 되므로
    # 첫 데이터행(data_start)의 서식을 이후 행에 복사 (%·천단위·테두리 유지)
    for i, kw in enumerate(keywords[:50]):
        r = data_start + i
        if r != data_start:
            _copy_row_style(ws, data_start, r, range(2, 15))
        ws.cell(row=r, column=3, value=kw.get("campaign", ""))        # C: 캠페인
        ws.cell(row=r, column=4, value=kw.get("adgroup", ""))         # D: 광고그룹
        ws.cell(row=r, column=5, value=kw.get("keyword", ""))         # E: 키워드
        ws.cell(row=r, column=6, value=kw.get("avg_rank", 0))         # F: 평균노출순위
        ws.cell(row=r, column=7, value=kw.get("impressions", 0) or 0)  # G: 노출
        ws.cell(row=r, column=8, value=kw.get("clicks", 0) or 0)       # H: 클릭
        ws.cell(row=r, column=9, value=f"=IFERROR(H{r}/G{r},0)")       # I: CTR
        ws.cell(row=r, column=10, value=f"=IFERROR(K{r}/H{r},0)")      # J: CPC
        ws.cell(row=r, column=11, value=kw.get("spend", 0) or 0)       # K: 광고비
        ws.cell(row=r, column=12, value=kw.get("conversions", 0) or 0)  # L: 전환수
        ws.cell(row=r, column=13, value=f"=IFERROR(L{r}/H{r},0)")      # M: CVR
        ws.cell(row=r, column=14, value=f"=IFERROR(K{r}/L{r},0)")      # N: 전환당비용

    # '기타(상위 외 키워드·확장검색 등)' 행 — 합계가 기간 총합과 정합되도록 잔차 채움
    shown = keywords[:50]
    n = len(shown)
    if period_total and n:
        etc_r = data_start + n
        disp_imp = sum(k.get("impressions", 0) or 0 for k in shown)
        disp_clk = sum(k.get("clicks", 0) or 0 for k in shown)
        disp_sp = sum(k.get("spend", 0) or 0 for k in shown)
        disp_cv = sum(k.get("conversions", 0) or 0 for k in shown)
        etc_imp = max(0, period_total.get("imp", 0) - disp_imp)
        etc_clk = max(0, period_total.get("clk", 0) - disp_clk)
        etc_sp = max(0, period_total.get("spend", 0) - disp_sp)
        etc_cv = max(0, period_total.get("conv", 0) - disp_cv)
        _copy_row_style(ws, data_start, etc_r, range(2, 15))
        ws.cell(row=etc_r, column=3, value="기타")                       # C
        ws.cell(row=etc_r, column=4, value="(상위 외 키워드·확장검색 등)")  # D
        ws.cell(row=etc_r, column=5, value="")                           # E
        ws.cell(row=etc_r, column=7, value=etc_imp)                      # G 노출
        ws.cell(row=etc_r, column=8, value=etc_clk)                      # H 클릭
        ws.cell(row=etc_r, column=9, value=f"=IFERROR(H{etc_r}/G{etc_r},0)")   # I CTR
        ws.cell(row=etc_r, column=10, value=f"=IFERROR(K{etc_r}/H{etc_r},0)")  # J CPC
        ws.cell(row=etc_r, column=11, value=etc_sp)                      # K 광고비
        ws.cell(row=etc_r, column=12, value=etc_cv)                      # L 전환수
        ws.cell(row=etc_r, column=13, value=f"=IFERROR(L{etc_r}/H{etc_r},0)")  # M CVR
        ws.cell(row=etc_r, column=14, value=f"=IFERROR(K{etc_r}/L{etc_r},0)")  # N 전환당비용


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  3. AI 코멘트 생성 (RAW 데이터 기반)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _aggregate_raw(raw_rows: list) -> dict:
    """RAW 행 리스트를 집계하여 코멘트용 요약 dict 생성.
    raw_rows: [[date_str, device, campaign, media, imp, clk, spend, conv], ...]
    """
    total = {"imp": 0, "clk": 0, "spend": 0, "conv": 0}
    by_device = {}   # device → {imp, clk, spend, conv}
    by_campaign = {}  # campaign → {imp, clk, spend, conv}
    by_week = {}     # week_num → {imp, clk, spend, conv}
    daily = {}       # date_str → {imp, clk, spend, conv}

    for row in raw_rows:
        date_str, device, campaign, media, imp, clk, spend, conv = row[:8]
        total["imp"] += imp
        total["clk"] += clk
        total["spend"] += spend
        total["conv"] += conv

        d = by_device.setdefault(device, {"imp": 0, "clk": 0, "spend": 0, "conv": 0})
        d["imp"] += imp; d["clk"] += clk; d["spend"] += spend; d["conv"] += conv

        c = by_campaign.setdefault(campaign, {"imp": 0, "clk": 0, "spend": 0, "conv": 0})
        c["imp"] += imp; c["clk"] += clk; c["spend"] += spend; c["conv"] += conv

        dd = daily.setdefault(date_str, {"imp": 0, "clk": 0, "spend": 0, "conv": 0})
        dd["imp"] += imp; dd["clk"] += clk; dd["spend"] += spend; dd["conv"] += conv

        # 주차 계산 (날짜에서 일 추출 → 7일 단위)
        try:
            day = int(date_str.replace(".", " ").split()[2])
            wk = (day - 1) // 7 + 1
            w = by_week.setdefault(wk, {"imp": 0, "clk": 0, "spend": 0, "conv": 0})
            w["imp"] += imp; w["clk"] += clk; w["spend"] += spend; w["conv"] += conv
        except (IndexError, ValueError):
            pass

    return {
        "total": total,
        "by_device": by_device,
        "by_campaign": by_campaign,
        "by_week": by_week,
        "daily": daily,
    }


def _build_v2_data_summary(
    agg: dict,
    keyword_data: list,
    client_name: str,
    year: int,
    month: int,
) -> str:
    """코멘트 생성용 데이터 요약 텍스트 작성."""
    t = agg["total"]
    imp, clk, spend, conv = t["imp"], t["clk"], t["spend"], t["conv"]
    ctr = round(clk / imp * 100, 2) if imp else 0
    cpc = round(spend / clk) if clk else 0
    cvr = round(conv / clk * 100, 2) if clk else 0
    cpa = round(spend / conv) if conv else 0

    lines = [
        f"클라이언트: {client_name}",
        f"보고 기간: {year}년 {month}월 (월간)",
        "",
        "■ 월 전체 실적",
        f"  노출수: {imp:,}회 / 클릭수: {clk:,}건 / CTR: {ctr}%",
        f"  CPC: {cpc:,}원 / 광고비: {spend:,}원",
        f"  전환수: {conv}건 / 전환율: {cvr}% / 전환당비용: {cpa:,}원",
    ]

    # 기기별
    lines.append("\n■ 기기별 실적")
    for dev in ("PC", "모바일", "전체"):
        d = agg["by_device"].get(dev)
        if not d or not d["imp"]:
            continue
        d_ctr = round(d["clk"] / d["imp"] * 100, 2) if d["imp"] else 0
        d_cvr = round(d["conv"] / d["clk"] * 100, 2) if d["clk"] else 0
        lines.append(
            f"  [{dev}] 노출 {d['imp']:,} / 클릭 {d['clk']:,} / CTR {d_ctr}% / "
            f"광고비 {d['spend']:,}원 / 전환 {d['conv']}건 / CVR {d_cvr}%"
        )

    # 주차별
    lines.append("\n■ 주차별 실적")
    for wk in sorted(agg["by_week"]):
        w = agg["by_week"][wk]
        w_ctr = round(w["clk"] / w["imp"] * 100, 2) if w["imp"] else 0
        lines.append(
            f"  {wk}주차: 노출 {w['imp']:,} / 클릭 {w['clk']:,} / CTR {w_ctr}% / "
            f"광고비 {w['spend']:,}원 / 전환 {w['conv']}건"
        )

    # 주차 간 비교
    weeks = sorted(agg["by_week"])
    if len(weeks) >= 2:
        last_wk = agg["by_week"][weeks[-1]]
        prev_wk = agg["by_week"][weeks[-2]]
        if prev_wk["clk"]:
            clk_chg = round((last_wk["clk"] / prev_wk["clk"] - 1) * 100, 1)
            lines.append(f"  → 전주 대비 클릭수 {clk_chg:+.1f}%")
        if prev_wk["conv"]:
            conv_chg = round((last_wk["conv"] / prev_wk["conv"] - 1) * 100, 1)
            lines.append(f"  → 전주 대비 전환수 {conv_chg:+.1f}%")

    # 캠페인별 상위 5
    top_camps = sorted(
        agg["by_campaign"].items(),
        key=lambda x: -(x[1]["spend"] or x[1]["imp"]),
    )[:5]
    if top_camps:
        lines.append("\n■ 캠페인별 실적 (광고비 상위 5)")
        for name, c in top_camps:
            c_ctr = round(c["clk"] / c["imp"] * 100, 2) if c["imp"] else 0
            lines.append(
                f"  [{name}] 노출 {c['imp']:,} / 클릭 {c['clk']:,} / CTR {c_ctr}% / "
                f"광고비 {c['spend']:,}원 / 전환 {c['conv']}건"
            )

    # 키워드 상위
    if keyword_data:
        lines.append("\n■ 키워드 상위 5 (노출 기준)")
        for kw in keyword_data[:5]:
            lines.append(
                f"  [{kw['keyword']}] 노출 {kw['impressions']:,} / 클릭 {kw['clicks']:,} / "
                f"전환 {kw.get('conversions', 0)}건"
            )

    return "\n".join(lines)


_V2_SYSTEM_PROMPT = """\
당신은 네이버 검색광고(파워링크)를 운영하는 디지털 마케팅 에이전시의 광고 운영 전문가입니다.
주간/월간 광고 성과 보고서의 '코멘트'를 작성합니다.

작성 규칙:
1. 첫 줄에 기간과 핵심 KPI 요약 (CTR / CPC / 전환수 / 전환율)
2. 주차별 또는 기기별(PC/MO) 성과 변화를 구체적 수치와 함께 서술
3. 성과 변동의 원인 분석 (키워드, 캠페인, 시즌, 운영 변경 등)
4. 전환 성과가 높은 키워드나 캠페인을 구체적으로 언급
5. 다음 기간 운영 방향 제시
6. 문체: 경어체(~하였습니다, ~예정입니다), 전문적이고 간결하게
7. 줄바꿈을 적절히 사용 (3~5개 단락, ㄴ 기호로 세부사항 기술)
8. 코멘트 텍스트만 출력 (설명·제목·따옴표 불필요)
9. 길이: 400~800자

참고 형식 예시:
6월 데이터
CTR 3.44% / CPC 6,047원 / 전환 수 29건 / 전환율 23.58%
 ㄴ 지난 5월 클릭 수 77건까지 감소, CTR 하락하며 저조한 성과 기록했으나, 이번 6월 클릭 수 123건, CTR 3.44% 기록하며 성과 회복
 ㄴ 평일, 주말 모두 클릭 및 전환 증가했으며 특히 평일에 CTR, CVR 큰 폭 상승
 ㄴ '핵심키워드' 키워드에서 가장 많은 전환 발생
"""


def _generate_v2_comment(
    raw_rows: list,
    keyword_data: list,
    client_name: str,
    year: int,
    month: int,
) -> str:
    """RAW 데이터를 집계하여 AI 코멘트를 생성."""
    import os
    import anthropic
    from dotenv import load_dotenv

    load_dotenv(str(_ROOT / ".env"), override=True)
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise RuntimeError(f"ANTHROPIC_API_KEY 미설정 (_ROOT={_ROOT})")

    agg = _aggregate_raw(raw_rows)
    data_summary = _build_v2_data_summary(agg, keyword_data, client_name, year, month)

    client = anthropic.Anthropic(api_key=api_key)
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        system=_V2_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": data_summary}],
    )
    return response.content[0].text.strip()


def _write_section_row(ws, r: int, vals: dict, has_group: bool = False):
    """캠페인별/그룹별 섹션의 한 행을 기입."""
    imp, clk, sp, cv = vals["imp"], vals["clk"], vals["spend"], vals["conv"]
    if has_group:
        ws.cell(row=r, column=2, value=vals.get("campaign", ""))
        ws.cell(row=r, column=3, value=vals.get("group", ""))
    else:
        ws.cell(row=r, column=2, value=vals.get("name", ""))
    ws.cell(row=r, column=4, value=imp)
    ws.cell(row=r, column=5, value=clk)
    ws.cell(row=r, column=6, value=round(clk / imp, 4) if imp else 0)
    ws.cell(row=r, column=7, value=round(sp / clk) if clk else 0)
    ws.cell(row=r, column=8, value=sp)
    ws.cell(row=r, column=9, value=cv)
    ws.cell(row=r, column=10, value=round(cv / clk, 4) if clk else 0)
    ws.cell(row=r, column=11, value=round(sp / cv) if cv else 0)


def _find_section(ws, header_keyword: str, search_start: int, search_end: int):
    """섹션 헤더·데이터시작·TOTAL 행 위치를 찾는다.
    Returns: (data_start_row, total_row, existing_data_rows)
    """
    header_row = None
    for r in range(search_start, search_end):
        val = ws.cell(row=r, column=2).value
        if val and isinstance(val, str) and header_keyword in val and "▶" in val:
            header_row = r
            break
    data_start = (header_row + 3) if header_row else search_start

    total_row = None
    for r in range(data_start, data_start + 200):
        val = ws.cell(row=r, column=2).value
        if val and isinstance(val, str) and "TOTAL" in val.upper():
            total_row = r
            break

    existing = (total_row - data_start) if total_row else 0
    return data_start, total_row, existing


def _copy_row_style(ws, src_row: int, dst_row: int, cols):
    """src_row의 셀 서식 + 행 높이를 dst_row로 복사 (초과행 무서식 보정)."""
    import copy as _copy
    for c in cols:
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst._style = _copy.copy(src._style)
    src_h = ws.row_dimensions[src_row].height
    if src_h is not None:
        ws.row_dimensions[dst_row].height = src_h


def _insert_rows_keep_merges(ws, idx: int, n: int):
    """insert_rows + 병합셀 이동 (openpyxl insert_rows는 병합범위를 옮기지 않는 버그 보정).
    idx 이상에서 시작하는 병합범위를 n행 아래로 재배치."""
    affected = [(m.min_row, m.min_col, m.max_row, m.max_col)
                for m in ws.merged_cells.ranges if m.min_row >= idx]
    for r1, c1, r2, c2 in affected:
        ws.unmerge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    ws.insert_rows(idx, n)
    for r1, c1, r2, c2 in affected:
        ws.merge_cells(start_row=r1 + n, start_column=c1, end_row=r2 + n, end_column=c2)


def _delete_rows_keep_merges(ws, idx: int, n: int):
    """delete_rows + 병합셀 이동. [idx, idx+n) 내부 병합은 제거, 그 아래는 n행 위로."""
    remove, shift = [], []
    for m in ws.merged_cells.ranges:
        coords = (m.min_row, m.min_col, m.max_row, m.max_col)
        if m.min_row >= idx + n:
            shift.append(coords)
        elif m.min_row >= idx:
            remove.append(coords)
    for r1, c1, r2, c2 in remove + shift:
        ws.unmerge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    ws.delete_rows(idx, n)
    for r1, c1, r2, c2 in shift:
        ws.merge_cells(start_row=r1 - n, start_column=c1, end_row=r2 - n, end_column=c2)


def _ensure_bc_merge(ws, first: int, last: int):
    """캠페인 블록 데이터행의 B:C 병합 보장(캠페인명이 B:C를 가로지름). 이미 병합된 행은 건너뜀."""
    merged = {m.min_row for m in ws.merged_cells.ranges
              if m.min_col == 2 and m.max_col >= 3 and m.min_row == m.max_row}
    for r in range(first, last + 1):
        if r not in merged:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)


def _resize_block(ws, data_start: int, total_row: int, need: int, style_src: int) -> int:
    """블록(data_start..total_row-1)을 need 행에 맞춰 insert/delete. 새 total_row 반환.
    insert 시 style_src 행의 서식을 새 행에 복사하고, 병합셀도 함께 이동시켜 정렬 유지."""
    cols = range(2, 12)
    existing = total_row - data_start
    if need > existing:
        n = need - existing
        _insert_rows_keep_merges(ws, total_row, n)
        for i in range(n):
            _copy_row_style(ws, style_src, total_row + i, cols)
        total_row += n
    elif need < existing:
        n = existing - need
        _delete_rows_keep_merges(ws, data_start + need, n)
        total_row -= n
    return total_row


def _clear_block(ws, first: int, last: int):
    for r in range(first, last + 1):
        for c in range(2, 12):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None


_RAW = "파워링크 RAW"


def _sumifs_camp(col: str, r: int) -> str:
    """캠페인블록: RAW를 매체='파워링크' + 캠페인(C열)=$B{r} 로 직접 집계."""
    return (f"=SUMIFS('{_RAW}'!{col}:{col},'{_RAW}'!$D:$D,\"파워링크\","
            f"'{_RAW}'!$C:$C,$B{r})")


def _write_campaign_row(ws, r: int, cname: str):
    """캠페인블록 한 행: B=캠페인명, D/E/H/I=RAW 직접 SUMIFS, F/G/J/K=파생수식."""
    ws.cell(row=r, column=2, value=cname)                        # B 캠페인
    ws.cell(row=r, column=4, value=_sumifs_camp("E", r))         # D 노출 ← RAW E
    ws.cell(row=r, column=5, value=_sumifs_camp("F", r))         # E 클릭 ← RAW F
    ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/D{r},)")      # F CTR
    ws.cell(row=r, column=7, value=f"=IFERROR(H{r}/E{r},0)")     # G CPC
    ws.cell(row=r, column=8, value=_sumifs_camp("G", r))         # H 광고비 ← RAW G
    ws.cell(row=r, column=9, value=_sumifs_camp("H", r))         # I 전환수 ← RAW H
    ws.cell(row=r, column=10, value=f"=IFERROR(I{r}/E{r},0)")    # J CVR
    ws.cell(row=r, column=11, value=f"=IFERROR(H{r}/I{r},0)")    # K 전환당비용


def _write_group_row(ws, r: int, cname: str, agname: str, st: dict):
    """그룹블록 한 행: B/C/D/E/H/I=값, F/G/J/K=파생수식."""
    ws.cell(row=r, column=2, value=cname)                        # B 캠페인
    ws.cell(row=r, column=3, value=agname)                       # C 그룹
    ws.cell(row=r, column=4, value=st["imp"])                    # D 노출
    ws.cell(row=r, column=5, value=st["clk"])                    # E 클릭
    ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/D{r},)")      # F CTR
    ws.cell(row=r, column=7, value=f"=IFERROR(H{r}/E{r},0)")     # G CPC
    ws.cell(row=r, column=8, value=st["spend"])                  # H 광고비
    ws.cell(row=r, column=9, value=st["conv"])                   # I 전환수
    ws.cell(row=r, column=10, value=f"=IFERROR(I{r}/E{r},0)")    # J CVR
    ws.cell(row=r, column=11, value=f"=IFERROR(H{r}/I{r},0)")    # K 전환당비용


def _write_block_total(ws, t: int, first: int, last: int):
    """블록 TOTAL 행: 값열(D/E/H/I)은 SUM, 파생열(F/G/J/K)은 IFERROR (F51/G51 버그 교정 포함)."""
    ws.cell(row=t, column=2, value=" TOTAL")
    ws.cell(row=t, column=4, value=f"=SUM(D{first}:D{last})")
    ws.cell(row=t, column=5, value=f"=SUM(E{first}:E{last})")
    ws.cell(row=t, column=6, value=f"=IFERROR(E{t}/D{t},)")
    ws.cell(row=t, column=7, value=f"=IFERROR(H{t}/E{t},0)")
    ws.cell(row=t, column=8, value=f"=SUM(H{first}:H{last})")
    ws.cell(row=t, column=9, value=f"=SUM(I{first}:I{last})")
    ws.cell(row=t, column=10, value=f"=IFERROR(I{t}/E{t},0)")
    ws.cell(row=t, column=11, value=f"=IFERROR(H{t}/I{t},0)")


def _fill_detail_sheet(wb, sheet_name: str, raw_rows: list, creative_data: list):
    """파워링크_상세데이터 갱신 (작성자 의도 재현).
    · 기기블록(7-9): 템플릿 SUMIFS 수식 유지(건드리지 않음)
    · 캠페인블록: 실제 캠페인마다 RAW 직접 SUMIFS 수식으로 재생성(동적 행수)
      → 페이파란 하드코딩·H15/I15 참조버그 제거, TOTAL이 월합과 정합
    · 그룹블록: creative_data(캠페인×광고그룹) 집계를 값으로, 파생지표는 수식
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    # ── 캠페인별 (RAW 직접 SUMIFS, 광고비 내림차순) ──
    camp_spend = {}
    for row in raw_rows:
        _, _, campaign, _, imp, clk, spend, conv = row[:8]
        camp_spend[campaign] = camp_spend.get(campaign, 0) + spend
    camps = [c for c, _ in sorted(camp_spend.items(), key=lambda x: -x[1])]
    need_camp = max(1, len(camps))

    camp_start, camp_total, _ = _find_section(ws, "캠페인", 10, 30)   # 14, 18
    camp_total = _resize_block(ws, camp_start, camp_total, need_camp, camp_start)
    _clear_block(ws, camp_start, camp_total - 1)
    for i, cname in enumerate(camps):
        _write_campaign_row(ws, camp_start + i, cname)
    _write_block_total(ws, camp_total, camp_start, camp_total - 1)
    # 캠페인명이 B:C를 가로지르도록 병합 유지(삽입행 포함)
    _ensure_bc_merge(ws, camp_start, camp_total)

    # ── 그룹별 (creative_data 캠페인×광고그룹, 값 기입) ──
    group_stats = {}
    for cr in creative_data:
        key = (cr.get("campaign", ""), cr.get("adgroup", ""))
        g = group_stats.setdefault(key, {"imp": 0, "clk": 0, "spend": 0, "conv": 0})
        g["imp"] += cr.get("impressions", 0) or 0
        g["clk"] += cr.get("clicks", 0) or 0
        g["spend"] += cr.get("spend", 0) or 0
        g["conv"] += cr.get("conversions", 0) or 0
    groups = sorted(group_stats.items(), key=lambda x: -x[1]["spend"])
    need_grp = max(1, len(groups))

    # 캠페인 블록 리사이즈로 그룹 헤더가 이동했으므로 camp_total 아래에서 다시 탐색
    grp_start, grp_total, _ = _find_section(ws, "그룹", camp_total + 1, camp_total + 40)
    grp_total = _resize_block(ws, grp_start, grp_total, need_grp, grp_start)
    _clear_block(ws, grp_start, grp_total - 1)
    for i, ((cname, agname), st) in enumerate(groups):
        _write_group_row(ws, grp_start + i, cname, agname, st)
    _write_block_total(ws, grp_total, grp_start, grp_total - 1)


def _write_comment_to_summary(wb, sheet_name: str, comment: str):
    """Summary 시트의 코멘트 셀(L9)에 텍스트 기입."""
    if sheet_name not in wb.sheetnames or not comment:
        return
    ws = wb[sheet_name]
    from openpyxl.styles import Alignment

    # L9 (row=9, col=12) — 원본 템플릿의 코멘트 위치
    cell = ws.cell(row=9, column=12)
    if not isinstance(cell, MergedCell):
        cell.value = comment
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  4. 메인 생성 함수
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def generate_v2_report(
    customer_id: str,
    year: int,
    month: int,
    client_name: str = "",
    output_dir: str | Path = None,
    out_name: str = None,
) -> Path:
    """버전B 보고서 생성 메인 함수.
    미래아이엔씨 양식 템플릿의 RAW 시트에 API 데이터를 채우고,
    날짜 참조를 갱신한 뒤 저장한다.
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    if not V2_TEMPLATE.exists():
        raise FileNotFoundError(f"V2 템플릿 파일 없음: {V2_TEMPLATE}")

    # ── 1. 데이터 수집 (RAW / 소재 / 키워드 평일·주말) ──
    if MOCK_MODE:
        raw_rows = _mock_raw_powerlink(year, month)
        creative_data = _mock_creative(year, month)
        keyword_wd = _mock_keyword_data(year, month)
        keyword_we = _mock_keyword_data(year, month)
    else:
        raw_rows, creative_data = [], []
        keyword_wd, keyword_we = [], []
        for cid in [c for c in str(customer_id).split("+") if c]:
            try:
                rows, camp_ids, camp_name = _collect_raw_powerlink(cid, year, month)
            except Exception:
                continue
            raw_rows.extend(rows)

            # 캠페인을 광고비 내림차순으로 정렬 → 키워드/소재 수집이 상위 캠페인 우선
            spend_by_name = {}
            for row in rows:
                spend_by_name[row[2]] = spend_by_name.get(row[2], 0) + row[6]
            camp_ids_ranked = sorted(
                camp_ids, key=lambda c: -spend_by_name.get(camp_name.get(c, ""), 0))

            last_day = calendar.monthrange(year, month)[1]
            tr = {"since": f"{year}-{month:02d}-01",
                  "until": f"{year}-{month:02d}-{last_day:02d}"}
            try:
                creative_data.extend(
                    _collect_raw_creative(cid, camp_ids_ranked[:20], camp_name, tr))
            except Exception:
                pass
            try:
                wd, we = _collect_keywords_split(
                    cid, camp_ids_ranked[:25], camp_name, year, month)
                keyword_wd.extend(wd)
                keyword_we.extend(we)
            except Exception:
                pass
        if not raw_rows:
            raise RuntimeError(f"데이터 수집 실패 (customer={customer_id})")
        creative_data.sort(key=lambda x: -(x.get("spend", 0) or 0))
        keyword_wd.sort(key=lambda x: -(x.get("spend", 0) or 0))
        keyword_we.sort(key=lambda x: -(x.get("spend", 0) or 0))

    # 코멘트용 월간 키워드 = 평일+주말 병합
    keyword_month = _merge_keywords(keyword_wd, keyword_we)

    # ── 템플릿 복사 ──
    safe_name = client_name or "클라이언트"
    if not out_name:
        out_name = f"{year}년{month:02d}월_{safe_name}_보고서B.xlsx"
    # 파일명 안전화(경로 구분자 제거)
    out_name = out_name.replace("/", "_").replace("\\", "_")
    out_path = output_dir / out_name
    shutil.copy2(V2_TEMPLATE, out_path)

    # ── 2. 엑셀 채우기 ──
    wb = openpyxl.load_workbook(str(out_path))

    # (a) 값-원천 시트: 파워링크 RAW + 파링_소재 RAW
    _fill_raw_sheet(wb, "파워링크 RAW", raw_rows)
    if creative_data:
        _fill_creative_raw(wb, "파링_소재 RAW", creative_data)

    # (b) Summary 자동계산 활성화: 일별 날짜 매칭 + 기간/주차 재생성
    _update_summary_dates(wb, "파워링크_Summary", year, month)
    _update_period_text(wb, year, month)

    # (c) 값+수식 시트: 상세데이터 · 소재별_주간 · 키워드(평일/주말)
    _fill_detail_sheet(wb, "파워링크_상세데이터", raw_rows, creative_data)
    if creative_data:
        _fill_creative_weekly(wb, "파워링크_소재별_주간", creative_data)
    # 키워드 합계를 기간 캠페인 총합과 정합시키기 위한 평일/주말 총합
    wd_total, we_total = _split_period_totals(raw_rows, year, month)
    if keyword_wd:
        _fill_keyword_sheet(wb, "파워링크_키워드_평일", keyword_wd, wd_total)
    if keyword_we:
        _fill_keyword_sheet(wb, "파워링크_키워드_주말", keyword_we, we_total)

    # (d) Excel이 파일을 열 때 전체 수식을 재계산하도록 강제
    wb.calculation.fullCalcOnLoad = True

    # (e) RAW+키워드 집계 기반 AI 코멘트 → Summary L9
    try:
        comment = _generate_v2_comment(raw_rows, keyword_month, safe_name, year, month)
    except Exception as e:
        comment = f"[코멘트 생성 실패] {e}"
    _write_comment_to_summary(wb, "파워링크_Summary", comment)

    wb.save(str(out_path))
    return out_path
