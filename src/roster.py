"""광고주 로스터(보관소) 로더 + accounts.json 동기화.

SoT(원본) 우선순위:
    1) Google Sheets  — ROSTER_GSHEET_ID 환경변수 + 서비스계정 크리덴셜이 있으면 사용
    2) 로컬 roster.xlsx — 위가 없으면 폴백 (설정 없이 바로 사용 가능)

로스터 컬럼(한글 헤더 기준, 영문 별칭도 허용):
    광고주명 | customer_id | api_key | api_secret | 매체 | 담당마케터 | 수신이메일 | 활성 | 그룹 | 대행사

- 로스터를 읽어 accounts.json(수집용 키 세트)으로 내려쓴다. accounts.json 은
  네이버 수집기(naver_searchad._load_accounts)가 읽는 파일이라 스키마를 맞춘다.
- 담당마케터/수신이메일/활성 같은 운영 메타는 roster 원본에만 두고 배치가 참조한다.

⚠️ accounts.json 은 시크릿을 담으므로 repo/외부에 노출 금지(로컬 전용).
"""
from __future__ import annotations

import json
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook

_ROOT = Path(__file__).resolve().parents[1]   # ad-report/
ROSTER_XLSX = _ROOT / "roster.xlsx"
ACCOUNTS_JSON = _ROOT / "accounts.json"

# 한글 헤더 → 표준 필드. 영문/변형 별칭도 매핑.
_HEADER_ALIASES = {
    "광고주명": "name", "광고주": "name", "name": "name", "이름": "name",
    "customer_id": "customer_id", "customerid": "customer_id",
    "커스터머아이디": "customer_id", "고객id": "customer_id", "cid": "customer_id",
    "api_key": "api_key", "apikey": "api_key", "액세스라이선스": "api_key", "키": "api_key",
    "api_secret": "api_secret", "apisecret": "api_secret",
    "시크릿": "api_secret", "비밀키": "api_secret",
    "매체": "media", "media": "media",
    "담당마케터": "marketer", "마케터": "marketer", "marketer": "marketer",
    "수신이메일": "email", "이메일": "email", "email": "email", "메일": "email",
    "활성": "active", "active": "active", "사용": "active", "상태": "active",
    "그룹": "group", "group": "group",
    "대행사": "agency", "agency": "agency",
}

_ACCOUNT_FIELDS = ("customer_id", "api_key", "api_secret", "name", "group", "media", "agency")


def _norm_header(h) -> str:
    return _HEADER_ALIASES.get(str(h or "").strip().lower().replace(" ", ""),
                               _HEADER_ALIASES.get(str(h or "").strip(), ""))


def _truthy(v) -> bool:
    """활성 컬럼 파싱. 빈칸/누락은 활성(True)으로 간주(온보딩 마찰 최소)."""
    if v is None or str(v).strip() == "":
        return True
    return str(v).strip().lower() in ("1", "y", "yes", "true", "o", "on", "활성", "사용", "예")


def _rows_to_records(header_row, data_rows) -> list[dict]:
    cols = [_norm_header(h) for h in header_row]
    records: list[dict] = []
    for row in data_rows:
        rec: dict = {}
        for key, val in zip(cols, row):
            if not key:
                continue
            rec[key] = ("" if val is None else str(val).strip())
        cid = rec.get("customer_id", "").strip()
        if not cid:
            continue   # customer_id 없는 행(빈 줄/주석) 스킵
        rec["customer_id"] = cid
        rec["active"] = _truthy(rec.get("active"))
        rec["agency"] = _truthy(rec.get("agency")) if rec.get("agency") not in (None, "") else False
        records.append(rec)
    return records


# ── 소스별 리더 ────────────────────────────────────────────
def _read_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    return _rows_to_records(rows[0], rows[1:])


def _read_csv(path: Path) -> list[dict]:
    import csv
    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return []
    return _rows_to_records(rows[0], rows[1:])


def read_file(path) -> list[dict]:
    """업로드된 로스터 파일(xlsx/csv)을 표준 레코드 리스트로 파싱."""
    path = Path(path)
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    if ext == ".csv":
        return _read_csv(path)
    raise ValueError(f"지원하지 않는 형식: {ext} (xlsx 또는 csv 만 가능)")


def _read_gsheet(sheet_id: str, worksheet: str | None) -> list[dict]:
    """gspread + 서비스계정으로 Google Sheet 읽기.
    필요: pip install gspread google-auth
    환경변수: GOOGLE_APPLICATION_CREDENTIALS = 서비스계정 json 경로
    시트는 서비스계정 이메일에 '뷰어' 공유되어 있어야 함."""
    import gspread   # 지연 import (미설치여도 xlsx 폴백 동작)

    cred_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "")
    if cred_path:
        gc = gspread.service_account(filename=cred_path)
    else:
        gc = gspread.service_account()   # 기본 위치 탐색
    sh = gc.open_by_key(sheet_id)
    ws = sh.worksheet(worksheet) if worksheet else sh.sheet1
    values = ws.get_all_values()
    if not values:
        return []
    return _rows_to_records(values[0], values[1:])


# ── 공개 API ───────────────────────────────────────────────
def load_roster() -> list[dict]:
    """SoT에서 로스터 로드. Google Sheet 우선, 없으면 로컬 roster.xlsx."""
    sheet_id = os.getenv("ROSTER_GSHEET_ID", "").strip()
    if sheet_id:
        return _read_gsheet(sheet_id, os.getenv("ROSTER_GSHEET_TAB", "").strip() or None)
    if ROSTER_XLSX.exists():
        return _read_xlsx(ROSTER_XLSX)
    raise FileNotFoundError(
        f"로스터 소스 없음: ROSTER_GSHEET_ID 환경변수도, {ROSTER_XLSX} 파일도 없습니다. "
        f"`python -m src.roster template` 로 템플릿을 먼저 만드세요."
    )


def sync_accounts_json(roster: list[dict] | None = None, preserve_extra: bool = True) -> Path:
    """로스터 → accounts.json (수집기가 읽는 키 세트) 내려쓰기.
    활성 계정만, 키/시크릿이 채워진 행만 기록한다.

    preserve_extra=True: 로스터에 없는 기존 accounts.json 계정(예: 에듀윌)은
    삭제하지 않고 유지한다(실수로 인한 계정 유실 방지). 같은 customer_id는 로스터가 우선.
    """
    roster = roster if roster is not None else load_roster()

    from_roster: dict[str, dict] = {}
    skipped = []
    for r in roster:
        if not r.get("active", True):
            continue
        if not (r.get("api_key") and r.get("api_secret")):
            skipped.append(r.get("name") or r["customer_id"])
            continue
        from_roster[r["customer_id"]] = {k: r.get(k, "") for k in _ACCOUNT_FIELDS}

    merged: dict[str, dict] = {}
    if preserve_extra and ACCOUNTS_JSON.exists():
        try:
            for a in json.loads(ACCOUNTS_JSON.read_text(encoding="utf-8")):
                cid = str(a.get("customer_id") or "").strip()
                if cid:
                    merged[cid] = a
        except Exception:
            pass
    merged.update(from_roster)   # 로스터가 기존값을 덮어씀(같은 cid)

    accounts = list(merged.values())
    ACCOUNTS_JSON.write_text(
        json.dumps(accounts, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    kept_extra = len(accounts) - len(from_roster)
    if skipped:
        print(f"⚠️  키/시크릿 미입력으로 제외된 계정: {', '.join(skipped)}")
    print(f"✅ accounts.json 동기화: 로스터 {len(from_roster)}개"
          + (f" + 기존 유지 {kept_extra}개" if kept_extra > 0 else "")
          + f" → {ACCOUNTS_JSON}")
    return ACCOUNTS_JSON


# ── 템플릿 생성 ────────────────────────────────────────────
_TEMPLATE_HEADERS = ["광고주명", "customer_id", "api_key", "api_secret",
                     "매체", "담당마케터", "수신이메일", "활성", "그룹", "대행사"]

# 시드 행: 광고주명 + customer_id 만 넣고, api_key/secret 은 배포 후 UI/로컬 시트로 입력.
# (실 키를 코드에 넣으면 GitHub 등에 유출되므로 비워둔다)
_SEED_ROWS = [
    ["에듀윌(파워링크)",       "4273",    "", "", "파워링크", "", "", "Y", "에듀윌", ""],
    ["린드스트롬",            "1543098", "", "", "파워링크", "", "", "Y", "", ""],
    ["센텍코리아(알코스캔)",  "1520112", "", "", "파워링크", "", "", "Y", "", ""],
    ["봄온아카데미(아나운서)", "322622",  "", "", "파워링크", "", "", "Y", "", ""],
    ["봄온아카데미(쇼호스트)", "3146213", "", "", "파워링크", "", "", "Y", "", ""],
    ["프로디지",              "1662708", "", "", "파워링크", "", "", "Y", "", ""],
]


def create_template(path: Path = ROSTER_XLSX, overwrite: bool = False) -> Path:
    """신규 5개 고객이 채워진 roster.xlsx 템플릿 생성."""
    if path.exists() and not overwrite:
        raise FileExistsError(f"이미 존재: {path} (덮어쓰려면 overwrite=True)")
    wb = Workbook()
    ws = wb.active
    ws.title = "roster"
    ws.append(_TEMPLATE_HEADERS)
    for row in _SEED_ROWS:
        ws.append(row)
    # 열 너비 살짝
    widths = [22, 12, 40, 30, 10, 10, 22, 6, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    wb.save(path)
    print(f"✅ 로스터 템플릿 생성: {path}  (api_key/api_secret 칸을 채우세요)")
    return path


if __name__ == "__main__":
    import sys
    cmd = sys.argv[1] if len(sys.argv) > 1 else "sync"
    if cmd == "template":
        create_template(overwrite="--force" in sys.argv)
    elif cmd == "sync":
        sync_accounts_json()
    elif cmd == "show":
        for r in load_roster():
            flag = "●" if r.get("active") else "○"
            keyed = "🔑" if (r.get("api_key") and r.get("api_secret")) else "  "
            print(f"{flag}{keyed} {r['customer_id']:>9}  {r.get('name','')}")
    else:
        print("사용법: python -m src.roster [template|sync|show]")
