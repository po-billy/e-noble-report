"""
Excel 보고서 생성 모듈.
기존 보고서 파일을 템플릿으로 복사한 후 새 데이터를 채워 넣는다.
브랜드검색 계정은 템플릿 없이 신규 생성.
"""
import copy
import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from collectors.base import ClientReport


_OUTPUT_DIR = Path(__file__).parent.parent / "output"
_TEMPLATE_PATH = Path(__file__).parent.parent / "data" / "report_template.xlsx"


# ── 공통 템플릿 채우기 (표지/월별/토탈/매체별) ────────────────
def _header_colmap(ws, header_row: int, c0: int = 2, c1: int = 20) -> dict:
    """헤더 행을 스캔해 '지표 키 → 컬럼 인덱스' 매핑 (컬럼 순서가 파일마다 달라도 안전)."""
    m: dict[str, int] = {}
    for c in range(c0, c1 + 1):
        v = str(ws.cell(row=header_row, column=c).value or "").strip()
        if not v:
            continue
        u = v.upper()
        if "노출" in v:                         m.setdefault("impressions", c)
        elif "클릭률" in v or "CTR" in u:        m.setdefault("ctr", c)
        elif "클릭당" in v:                      m.setdefault("cpc", c)
        elif "클릭" in v:                        m.setdefault("clicks", c)
        elif "광고비" in v:                      m.setdefault("spend", c)
        elif "평균순위" in v or "순위" in v:      m.setdefault("avg_rank", c)
        elif "전환매출" in v or "매출" in v:      m.setdefault("revenue", c)
        elif "전환율" in v:                      m.setdefault("conversion_rate", c)
        elif "전환당" in v:                      m.setdefault("cpa", c)
        elif "전환" in v:                        m.setdefault("conversions", c)
        elif "수익률" in v or "ROAS" in u:       m.setdefault("roas", c)
        elif "코멘트" in v:                      m.setdefault("comment", c)
        elif "월별" in v or "유형" in v or "캠페인" in v: m.setdefault("label", c)
    return m


_STAT_KEYS = ("impressions", "clicks", "ctr", "cpc", "spend", "avg_rank",
              "conversions", "revenue", "conversion_rate", "cpa", "roas")


def _write_row(ws, row: int, colmap: dict, label: str, stat: dict):
    """colmap 에 맞춰 한 행을 기입 (비율 지표는 0~1 비율 그대로 — 셀 % 서식이 ×100 표시)."""
    if "label" in colmap:
        ws.cell(row=row, column=colmap["label"]).value = label
    for key in _STAT_KEYS:
        if key in colmap and key in stat:
            ws.cell(row=row, column=colmap[key]).value = stat.get(key, 0)


def _fill_monthly_sheet(ws, history: list, comment: str):
    """월별/토탈 시트: 헤더 아래 데이터 영역에 월별 행을 하단 정렬로 기입."""
    # 헤더 행 탐색 (월별 + 노출수)
    header_row = None
    for r in range(1, min(ws.max_row, 40) + 1):
        vals = [str(ws.cell(row=r, column=c).value or "") for c in range(2, 16)]
        if any("월별" in v for v in vals) and any("노출" in v for v in vals):
            header_row = r
            break
    if header_row is None:
        return
    colmap = _header_colmap(ws, header_row)

    # 데이터 영역 마지막 행 = 코멘트 병합블록(N{header+1}:R{last})의 끝, 없으면 header+13
    last_row = header_row + 13
    for mr in ws.merged_cells.ranges:
        if mr.min_col >= colmap.get("comment", 99) and mr.min_row > header_row:
            last_row = max(last_row, mr.max_row)

    hist = history[-13:]                       # 안전 상한
    start = max(header_row + 1, last_row - len(hist) + 1)
    for i, h in enumerate(hist):
        _write_row(ws, start + i, colmap, f"{h.get('month')}월", h)

    # 코멘트는 병합블록 좌상단(헤더+1행)에 기입
    if comment and "comment" in colmap:
        ws.cell(row=header_row + 1, column=colmap["comment"]).value = comment


def _fill_media_sheet(ws, media_breakdown: list):
    """매체별 시트: 합계 + 매체유형별(파워링크/브랜드검색 등) 행 기입."""
    import re
    header_row = None
    for r in range(1, min(ws.max_row, 20) + 1):
        vals = [str(ws.cell(row=r, column=c).value or "") for c in range(2, 16)]
        if any("유형" in v for v in vals) and any("노출" in v for v in vals):
            header_row = r
            break
    if header_row is None:
        return
    colmap = _header_colmap(ws, header_row)

    # [태그] 캠페인명 → 매체유형별 합산
    groups: dict[str, dict] = {}
    for m in media_breakdown:
        tag = re.match(r"\[(.+?)\]", m.get("media_name", ""))
        key = tag.group(1) if tag else "기타"
        g = groups.setdefault(key, {"impressions": 0, "clicks": 0, "spend": 0,
                                    "conversions": 0, "revenue": 0, "_rankw": 0.0})
        for f in ("impressions", "clicks", "spend", "conversions", "revenue"):
            g[f] += m.get(f, 0) or 0
        g["_rankw"] += (m.get("avg_rank", 0) or 0) * (m.get("impressions", 0) or 0)

    def _derive(g):
        imp, clk, sp, cv, rv = g["impressions"], g["clicks"], g["spend"], g["conversions"], g["revenue"]
        return {
            "impressions": imp, "clicks": clk,
            "ctr": round(clk / imp, 4) if imp else 0,
            "cpc": round(sp / clk) if clk else 0,
            "spend": sp,
            "avg_rank": round(g["_rankw"] / imp, 1) if imp else 0,
            "conversions": cv,
            "conversion_rate": round(cv / clk, 4) if clk else 0,
            "revenue": rv,
            "cpa": round(sp / cv) if cv else 0,
            "roas": round(rv / sp, 2) if sp else 0,
        }

    ordered = sorted(groups.items(), key=lambda kv: -(kv[1]["spend"] or kv[1]["impressions"] or 0))
    total = {"impressions": 0, "clicks": 0, "spend": 0, "conversions": 0, "revenue": 0, "_rankw": 0.0}
    for _k, g in ordered:
        for f in total:
            total[f] += g[f]

    r = header_row + 1
    _write_row(ws, r, colmap, "합계", _derive(total)); r += 1
    for tag, g in ordered:
        _write_row(ws, r, colmap, tag, _derive(g)); r += 1


def generate_templated_report(report: ClientReport, raw_data: dict,
                              output_dir=None, template_path=None) -> Path:
    """공통 템플릿(data/report_template.xlsx)에 검색광고+브랜드검색 합산 데이터를 채워 보고서 생성.
    시트: 표지 · 월별 · 토탈(브검포함) · 매체별 (상세 시트는 템플릿에 없음)."""
    out_dir = Path(output_dir) if output_dir else _OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    tpl = Path(template_path) if template_path else _TEMPLATE_PATH

    fname = f"{report.year}년{report.month}월_{report.client_name}_월간보고서.xlsx"
    out_path = out_dir / fname
    shutil.copy2(str(tpl), str(out_path))

    import warnings as _w
    with _w.catch_warnings():
        _w.simplefilter("ignore")
        wb = openpyxl.load_workbook(str(out_path))

    history = raw_data.get("monthly_history", [])
    media   = raw_data.get("media_breakdown", [])

    # 표지 제목
    if "표지" in wb.sheetnames:
        ws = wb["표지"]
        for r in range(9, 14):
            cell = ws.cell(row=r, column=1)
            if cell.value and "데이터" in str(cell.value):
                cell.value = f"{report.client_name} {report.year}년 {report.month}월 데이터"
                break
        else:
            ws["A11"] = f"{report.client_name} {report.year}년 {report.month}월 데이터"

    # 월별 / 토탈(브검포함) — 검색광고+브랜드검색 합산 동일 기입
    for sname in ("월별", "토탈(브검포함)"):
        if sname in wb.sheetnames:
            _fill_monthly_sheet(wb[sname], history, report.comment)

    # 매체별 — 파워링크/브랜드검색 등 유형별 분해
    if "매체별" in wb.sheetnames:
        _fill_media_sheet(wb["매체별"], media)

    wb.save(str(out_path))
    return out_path


def _find_cell(ws, text: str):
    """시트에서 text를 포함하는 첫 번째 셀 반환"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and text in str(cell.value):
                return cell
    return None


def _update_cover(ws, report: ClientReport):
    """표지 시트 업데이트: 제목, 운영기간"""
    # 제목 행 (row 11 부근) 업데이트
    for row in ws.iter_rows(min_row=9, max_row=13):
        for cell in row:
            v = str(cell.value or "").strip()
            if report.client_name in v or ("월" in v and "보고서" in v):
                cell.value = f"{report.client_name} {report.year}년 {report.month}월 데이터"
                break

    # 운영기간 업데이트
    cell = _find_cell(ws, "운영기간")
    if cell:
        # 같은 행의 다음 값 셀
        period_cell = ws.cell(row=cell.row, column=cell.column + 1)
        last_day = _last_day_of_month(report.year, report.month)
        period_cell.value = (
            f"{report.year}-{report.month:02d}-01 ~ "
            f"{report.year}-{report.month:02d}-{last_day:02d}"
        )


def _last_day_of_month(year: int, month: int) -> int:
    import calendar
    return calendar.monthrange(year, month)[1]


def _find_monthly_header_row(ws) -> int | None:
    """월별 시트에서 '월별', '노출수' 헤더가 있는 행 번호 반환"""
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip() if c else "" for c in row]
        if "월별" in cells and "노출수" in cells and "클릭수" in cells:
            return i
    return None


def _find_monthly_data_rows(ws, header_row: int) -> list[int]:
    """헤더 행 아래의 실제 데이터 행 번호 목록 반환"""
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(c is not None for c in row):
            continue
        # 월 레이블이 있거나 숫자 데이터가 있는 행
        if row[1] is not None and any(isinstance(c, (int, float)) for c in row[2:8]):
            rows.append(i)
    return rows


def _update_monthly_stats(ws, report: ClientReport):
    """월별 시트에 실제 수치(노출/클릭/전환 등)를 최신 월 행에 기입"""
    header_row = _find_monthly_header_row(ws)
    if header_row is None:
        return

    headers = [str(c).strip() if c else "" for c in
               next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]

    # 컬럼 인덱스 매핑 (1-indexed)
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers, start=1):
        lh = h.lower()
        if "노출" in lh:                           col_map.setdefault("impressions", i)
        elif "클릭" in lh and "률" not in lh:     col_map.setdefault("clicks", i)
        elif "클릭률" in lh or "ctr" in lh:       col_map.setdefault("ctr", i)
        elif "전환매출" in lh or "매출액" in lh:  col_map.setdefault("revenue", i)
        elif "전환율" in lh:                       col_map.setdefault("conversion_rate", i)
        elif "전환수" in lh or "전환건" in lh:    col_map.setdefault("conversions", i)
        elif "광고비" in lh or "비용" in lh:      col_map.setdefault("spend", i)
        elif "roas" in lh:                         col_map.setdefault("roas", i)

    if not col_map:
        return

    mt = report.monthly_total
    # ctr, conversion_rate는 0~1 비율 그대로 저장.
    # 기존 템플릿 셀에 "0.00%" 포맷이 있으면 Excel이 알아서 ×100 표시.
    # 포맷이 없는 셀은 raw 비율로 보이지만 보고서 생성 후 수동 포맷 가능.
    stat_values = {
        "impressions":     mt.impressions,
        "clicks":          mt.clicks,
        "ctr":             mt.ctr if mt.ctr <= 1 else mt.ctr / 100,
        "spend":           mt.spend,
        "conversions":     mt.conversions,
        "conversion_rate": mt.conversion_rate if mt.conversion_rate <= 1 else mt.conversion_rate / 100,
        "revenue":         mt.revenue,
        "roas":            mt.roas,
    }

    # 현재 월 레이블(예: "2026.05", "5월", "26년5월" 등)에 해당하는 행 찾기
    month_patterns = [
        f"{report.year}.{report.month:02d}",
        f"{report.year}년 {report.month}월",
        f"{report.month}월",
        f"{str(report.year)[2:]}년{report.month}월",
    ]

    target_row = None
    # 가장 마지막 데이터 행을 기본값으로
    data_rows = _find_monthly_data_rows(ws, header_row)
    if data_rows:
        target_row = data_rows[-1]

    # 월 레이블이 일치하는 행 우선 탐색
    for row_idx, row_vals in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        row_str = " ".join(str(c) for c in row_vals if c is not None)
        if any(p in row_str for p in month_patterns):
            target_row = row_idx
            break

    if target_row is None:
        return

    for key, col_idx in col_map.items():
        val = stat_values.get(key)
        if val is not None:
            ws.cell(row=target_row, column=col_idx).value = val


def _update_monthly_comment(ws, report: ClientReport, new_comment: str):
    """월별 시트에서 최신 월 행의 코멘트 셀 업데이트"""
    header_row = _find_monthly_header_row(ws)
    if header_row is None:
        return

    headers = [str(c).strip() if c else "" for c in
               next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]

    # '코멘트' 컬럼 인덱스 찾기 (1-indexed)
    comment_col = None
    for i, h in enumerate(headers, start=1):
        if "코멘트" in h:
            comment_col = i
            break

    if comment_col is None:
        return

    # 기존 코멘트가 있는 셀 찾기 (코멘트 컬럼에서 가장 긴 텍스트)
    best_row = None
    best_len = 0
    for row in ws.iter_rows(min_row=header_row + 1):
        cell = row[comment_col - 1]
        if cell.value and isinstance(cell.value, str) and len(cell.value) > best_len:
            best_len = len(cell.value)
            best_row = cell.row

    if best_row:
        target_cell = ws.cell(row=best_row, column=comment_col)
        target_cell.value = new_comment
        target_cell.alignment = Alignment(wrap_text=True, vertical="top")
    else:
        # 코멘트 셀이 없으면 데이터 행 중 가장 마지막 행에 삽입
        data_rows = _find_monthly_data_rows(ws, header_row)
        if data_rows:
            ws.cell(row=data_rows[-1], column=comment_col).value = new_comment


def generate_report(
    report: ClientReport,
    template_path: str | Path,
    output_dir: str | Path | None = None,
) -> Path:
    """
    template_path를 복사하고 report 데이터로 업데이트한 후 파일 경로 반환.
    """
    out_dir = Path(output_dir) if output_dir else _OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    fname = f"{report.year}년{report.month}월_{report.client_name}_월간보고서.xlsx"
    out_path = out_dir / fname

    shutil.copy2(str(template_path), str(out_path))

    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(str(out_path))

    if "표지" in wb.sheetnames:
        _update_cover(wb["표지"], report)

    for sname in ("월별", "토탈(브검포함)"):
        if sname in wb.sheetnames:
            _update_monthly_stats(wb[sname], report)
            _update_monthly_comment(wb[sname], report, report.comment)
            break

    wb.save(str(out_path))
    return out_path


# ── 브랜드검색 전용 Excel 생성 (템플릿 없이 신규 생성) ──────────

_HEADER_FILL  = PatternFill("solid", fgColor="2F4F8F")
_HEADER_FONT  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
_DATA_FONT    = Font(name="맑은 고딕", size=10)
_TITLE_FONT   = Font(name="맑은 고딕", bold=True, size=14)
_SUB_FONT     = Font(name="맑은 고딕", bold=True, size=11)
_THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _HEADER_FONT
    c.fill = _HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _THIN_BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _val(ws, row, col, value, num_fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="맑은 고딕", size=10, bold=bold)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _THIN_BORDER
    if num_fmt:
        c.number_format = num_fmt
    return c


# 숫자 표시 포맷
_FMT_INT = "#,##0"
_FMT_WON = "₩#,##0"
_FMT_PCT = "0.00%"
_FMT_ROAS = "0%"
_FMT_RANK = "0.0"

_KPI_LABEL_FILL = PatternFill("solid", fgColor="EEF2FA")
_DELTA_UP_FONT   = Font(name="맑은 고딕", size=9, color="C0392B")   # 증가(빨강)
_DELTA_DOWN_FONT = Font(name="맑은 고딕", size=9, color="1F6FB2")   # 감소(파랑)


def _pct_ratio(v) -> float:
    """비율(0~1) 보장. API가 %(예: 2.5)로 줄 수도 있어 방어."""
    v = v or 0
    return v if v <= 1 else v / 100


def _kpi_card(ws, row, col, label, value, fmt, delta_pct=None):
    """KPI 카드 한 칸: 위=라벨, 아래=값(+전월대비)"""
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(name="맑은 고딕", bold=True, size=9, color="555555")
    lc.fill = _KPI_LABEL_FILL
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = _THIN_BORDER

    vc = ws.cell(row=row + 1, column=col, value=value)
    vc.font = Font(name="맑은 고딕", bold=True, size=13)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = _THIN_BORDER
    if fmt:
        vc.number_format = fmt

    dc = ws.cell(row=row + 2, column=col)
    if delta_pct is not None:
        arrow = "▲" if delta_pct > 0 else ("▼" if delta_pct < 0 else "―")
        dc.value = f"{arrow} {abs(delta_pct):.1f}% (전월비)"
        dc.font = _DELTA_UP_FONT if delta_pct > 0 else (_DELTA_DOWN_FONT if delta_pct < 0 else _DATA_FONT)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    dc.border = _THIN_BORDER


def _delta(cur, prev) -> float | None:
    if not prev:
        return None
    return round((cur - prev) / prev * 100, 1)


def _table(ws, start_row, columns, rows_data, total_label=None):
    """헤더 + 데이터 + (옵션)합계 행을 그리는 범용 표.
    columns: [(헤더, key, fmt, width), ...]  / rows_data: [dict, ...]
    return: 마지막으로 쓴 행 번호"""
    for col, (h, _k, _f, w) in enumerate(columns, start=1):
        _hdr(ws, start_row, col, h, w)

    r = start_row
    for r, item in enumerate(rows_data, start=start_row + 1):
        for col, (_h, key, fmt, _w) in enumerate(columns, start=1):
            val = item.get(key, "")
            cell = _val(ws, r, col, val, fmt if isinstance(val, (int, float)) else None)
            if col == 1 and isinstance(val, str):
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 합계는 누적 가능한 지표만 (CPC·CPA·CTR·ROAS·순위 등 비율/단가 지표는 제외)
    _SUMMABLE = {"impressions", "clicks", "spend", "conversions", "revenue"}
    if total_label and rows_data:
        r += 1
        for col, (_h, key, fmt, _w) in enumerate(columns, start=1):
            if col == 1:
                _val(ws, r, col, total_label, bold=True)
            elif key in _SUMMABLE:
                _val(ws, r, col, sum(d.get(key, 0) or 0 for d in rows_data), fmt, bold=True)
            else:
                _val(ws, r, col, "", bold=True)
    return r


def generate_full_report(
    report: ClientReport,
    raw_data: dict,
    output_dir: str | Path | None = None,
) -> Path:
    """
    네이버 검색광고 통계로 종합 운영 보고서 Excel을 신규 생성한다 (조회 데이터 전용).
    raw_data: get_monthly_stats() 반환값
      {monthly_total, monthly_history, media_breakdown, daily_stats, keyword_stats}
    시트: 표지 · 핵심요약 · 월별추이 · 캠페인별 · 일별추이 · 키워드TOP
    """
    out_dir = Path(output_dir) if output_dir else _OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    mt        = raw_data.get("monthly_total", {})
    history   = raw_data.get("monthly_history", [])
    media_bkd = raw_data.get("media_breakdown", [])
    daily     = raw_data.get("daily_stats", [])
    keywords  = raw_data.get("keyword_stats", [])

    total_spend = sum(m.get("spend", 0) or 0 for m in media_bkd) or (mt.get("spend", 0) or 0)
    is_fixed_fee = total_spend == 0 and (mt.get("impressions", 0) or 0) > 0  # 브랜드검색(고정비)

    kind = "브랜드검색" if is_fixed_fee else "검색광고"
    fname = f"{report.year}년{report.month}월_{report.client_name}_{kind}_보고서.xlsx"
    out_path = out_dir / fname

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    last_day = _last_day_of_month(report.year, report.month)
    period = f"{report.year}-{report.month:02d}-01 ~ {report.year}-{report.month:02d}-{last_day:02d}"

    # ── Sheet 1: 표지 ────────────────────────────────────────
    ws = wb.create_sheet("표지")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 46
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = f"{report.client_name} {report.year}년 {report.month}월 운영보고서"
    t.font = _TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    info = [
        ("광고주", report.client_name),
        ("운영기간", period),
        ("매체", "네이버 브랜드검색" if is_fixed_fee else "네이버 검색광고"),
        ("담당자", report.manager or "-"),
        ("연락처", report.phone or "-"),
        ("이메일", report.email or "-"),
    ]
    for i, (label, value) in enumerate(info, start=3):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = Font(name="맑은 고딕", bold=True, size=10)
        lc.fill = _KPI_LABEL_FILL
        lc.border = _THIN_BORDER
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws.cell(row=i, column=2, value=value)
        vc.font = _DATA_FONT
        vc.border = _THIN_BORDER
        vc.alignment = Alignment(horizontal="left", vertical="center")

    # ── Sheet 2: 핵심요약 (KPI 대시보드) ─────────────────────
    ws2 = wb.create_sheet("핵심요약")
    for c in "ABCDE":
        ws2.column_dimensions[c].width = 18
    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value = f"핵심 성과 요약  ({report.year}년 {report.month}월)"
    t2.font = _TITLE_FONT
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 34

    prev = history[-2] if len(history) >= 2 else {}

    def d(key):  # 전월 대비 증감률
        return _delta(mt.get(key, 0) or 0, prev.get(key, 0) or 0)

    # KPI 카드: 3행 묶음(라벨/값/증감) × 행 배치
    cards = [
        ("노출수", mt.get("impressions", 0), _FMT_INT, d("impressions")),
        ("클릭수", mt.get("clicks", 0), _FMT_INT, d("clicks")),
        ("클릭률(CTR)", _pct_ratio(mt.get("ctr", 0)), _FMT_PCT, None),
        ("광고비", mt.get("spend", 0), _FMT_WON, d("spend")),
        ("전환수", mt.get("conversions", 0), _FMT_INT, d("conversions")),
        ("전환율", _pct_ratio(mt.get("conversion_rate", 0)), _FMT_PCT, None),
        ("전환매출", mt.get("revenue", 0), _FMT_WON, d("revenue")),
        ("ROAS", mt.get("roas", 0), _FMT_ROAS, d("roas")),
        ("CPC", mt.get("cpc", 0), _FMT_WON, None),
        ("CPA", mt.get("cpa", 0), _FMT_WON, None),
        ("평균순위", mt.get("avg_rank", 0), _FMT_RANK, None),
    ]
    PER_ROW = 5
    row = 3
    for i, (label, value, fmt, dp) in enumerate(cards):
        col = (i % PER_ROW) + 1
        if i and col == 1:
            row += 4
        _kpi_card(ws2, row, col, label, value, fmt, dp)
    note_row = row + 4
    ws2.merge_cells(f"A{note_row}:E{note_row}")
    note = ws2.cell(row=note_row, column=1)
    if is_fixed_fee:
        note.value = "※ 브랜드검색(고정비 계약) 계정 — CPC·ROAS 등 과금 지표는 참고용입니다."
    else:
        note.value = "※ 광고비는 salesAmt(집행액, VAT 별도) 기준입니다."
    note.font = Font(name="맑은 고딕", size=9, color="888888")

    if report.comment:
        cmt_row = note_row + 2
        ws2.cell(row=cmt_row, column=1, value="■ 코멘트").font = _SUB_FONT
        ws2.merge_cells(f"A{cmt_row+1}:E{cmt_row+5}")
        cc = ws2.cell(row=cmt_row + 1, column=1, value=report.comment)
        cc.font = _DATA_FONT
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        cc.border = _THIN_BORDER

    # ── Sheet 3: 월별 추이 ───────────────────────────────────
    ws3 = wb.create_sheet("월별추이")
    ws3.merge_cells("A1:J1")
    ws3["A1"].value = "월별 추이 (최근 3개월)"
    ws3["A1"].font = _SUB_FONT
    ws3.row_dimensions[1].height = 26

    month_cols = [
        ("월별", "_label", None, 12), ("노출수", "impressions", _FMT_INT, 14),
        ("클릭수", "clicks", _FMT_INT, 12), ("클릭률", "ctr", _FMT_PCT, 11),
        ("광고비", "spend", _FMT_WON, 15), ("전환수", "conversions", _FMT_INT, 11),
        ("전환율", "conversion_rate", _FMT_PCT, 11), ("전환매출", "revenue", _FMT_WON, 16),
        ("ROAS", "roas", _FMT_ROAS, 10), ("코멘트", "_comment", None, 48),
    ]
    hist_rows = []
    for h in history:
        is_cur = (h.get("year") == report.year and h.get("month") == report.month)
        hist_rows.append({
            "_label": f"{h.get('year')}년 {h.get('month')}월",
            "impressions": h.get("impressions", 0), "clicks": h.get("clicks", 0),
            "ctr": _pct_ratio(h.get("ctr", 0)), "spend": h.get("spend", 0),
            "conversions": h.get("conversions", 0), "conversion_rate": _pct_ratio(h.get("conversion_rate", 0)),
            "revenue": h.get("revenue", 0), "roas": h.get("roas", 0),
            "_comment": report.comment if is_cur else "",
        })
    end_r = _table(ws3, 2, month_cols, hist_rows)
    # 당월 행 강조 + 코멘트 줄바꿈
    cur_excel_row = 2 + len(hist_rows)
    for col in range(1, len(month_cols) + 1):
        ws3.cell(row=cur_excel_row, column=col).font = Font(name="맑은 고딕", size=10, bold=True)
    cmt_cell = ws3.cell(row=cur_excel_row, column=len(month_cols))
    cmt_cell.alignment = Alignment(wrap_text=True, vertical="top")
    if report.comment:
        ws3.row_dimensions[cur_excel_row].height = 90

    # 전월 대비 변화율
    if len(history) >= 2:
        rr = end_r + 2
        ws3.cell(row=rr, column=1, value="■ 전월 대비").font = _SUB_FONT
        for i in range(1, len(history)):
            cur, prv = history[i], history[i - 1]
            line = ws3.cell(
                row=rr + i, column=1,
                value=f"{prv.get('month')}월→{cur.get('month')}월",
            )
            line.font = _DATA_FONT
            for col, key in ((2, "impressions"), (3, "clicks"), (5, "spend"), (8, "revenue")):
                dv = _delta(cur.get(key, 0) or 0, prv.get(key, 0) or 0)
                cell = ws3.cell(row=rr + i, column=col, value=(f"{dv:+.1f}%" if dv is not None else "-"))
                cell.font = _DELTA_UP_FONT if (dv or 0) > 0 else (_DELTA_DOWN_FONT if (dv or 0) < 0 else _DATA_FONT)

    # ── Sheet 4: 캠페인별 ────────────────────────────────────
    ws4 = wb.create_sheet("캠페인별")
    ws4.merge_cells("A1:I1")
    ws4["A1"].value = f"캠페인별 실적 ({report.year}년 {report.month}월)"
    ws4["A1"].font = _SUB_FONT
    ws4.row_dimensions[1].height = 26
    camp_cols = [
        ("캠페인명", "media_name", None, 34), ("노출수", "impressions", _FMT_INT, 14),
        ("클릭수", "clicks", _FMT_INT, 12), ("클릭률", "ctr", _FMT_PCT, 11),
        ("광고비", "spend", _FMT_WON, 15), ("CPC", "cpc", _FMT_WON, 11),
        ("전환수", "conversions", _FMT_INT, 11), ("전환매출", "revenue", _FMT_WON, 16),
        ("ROAS", "roas", _FMT_ROAS, 10),
    ]
    camp_rows = [dict(m, ctr=_pct_ratio(m.get("ctr") or (m.get("clicks", 0) / m["impressions"] if m.get("impressions") else 0)))
                 for m in media_bkd]
    _table(ws4, 2, camp_cols, camp_rows, total_label="합계")

    # ── Sheet 5: 일별 추이 ───────────────────────────────────
    if daily:
        ws5 = wb.create_sheet("일별추이")
        ws5.merge_cells("A1:H1")
        ws5["A1"].value = f"일별 추이 ({report.year}년 {report.month}월)"
        ws5["A1"].font = _SUB_FONT
        ws5.row_dimensions[1].height = 26
        daily_cols = [
            ("날짜", "date", None, 13), ("요일", "weekday", None, 9),
            ("노출수", "impressions", _FMT_INT, 13), ("클릭수", "clicks", _FMT_INT, 11),
            ("클릭률", "ctr", _FMT_PCT, 11), ("광고비", "spend", _FMT_WON, 14),
            ("전환수", "conversions", _FMT_INT, 10), ("전환매출", "revenue", _FMT_WON, 15),
        ]
        daily_rows = []
        for dr in daily:
            imp = dr.get("impressions", 0) or 0
            daily_rows.append({
                "date": str(dr.get("date", "")), "weekday": dr.get("weekday", ""),
                "impressions": imp, "clicks": dr.get("clicks", 0),
                "ctr": _pct_ratio(dr.get("ctr") or (dr.get("clicks", 0) / imp if imp else 0)),
                "spend": dr.get("spend", 0), "conversions": dr.get("conversions", 0),
                "revenue": dr.get("revenue", 0),
            })
        _table(ws5, 2, daily_cols, daily_rows, total_label="합계")

    # ── Sheet 6: 키워드 TOP ──────────────────────────────────
    if keywords:
        ws6 = wb.create_sheet("키워드TOP")
        ws6.merge_cells("A1:J1")
        ws6["A1"].value = f"키워드 TOP {len(keywords)} ({report.year}년 {report.month}월 · 노출수순)"
        ws6["A1"].font = _SUB_FONT
        ws6.row_dimensions[1].height = 26
        kw_cols = [
            ("순위", "_rank", None, 7), ("키워드", "keyword", None, 26),
            ("노출수", "impressions", _FMT_INT, 13), ("클릭수", "clicks", _FMT_INT, 11),
            ("클릭률", "ctr", _FMT_PCT, 11), ("광고비", "spend", _FMT_WON, 14),
            ("CPC", "cpc", _FMT_WON, 11), ("전환수", "conversions", _FMT_INT, 10),
            ("전환매출", "revenue", _FMT_WON, 15), ("ROAS", "roas", _FMT_ROAS, 10),
        ]
        kw_rows = []
        for i, kw in enumerate(keywords, start=1):
            imp = kw.get("impressions", 0) or 0
            kw_rows.append({
                "_rank": i, "keyword": kw.get("keyword", ""),
                "impressions": imp, "clicks": kw.get("clicks", 0),
                "ctr": _pct_ratio(kw.get("ctr") or (kw.get("clicks", 0) / imp if imp else 0)),
                "spend": kw.get("spend", 0), "cpc": kw.get("cpc", 0),
                "conversions": kw.get("conversions", 0), "revenue": kw.get("revenue", 0),
                "roas": kw.get("roas", 0),
            })
        _table(ws6, 2, kw_cols, kw_rows, total_label="합계")

    wb.save(str(out_path))
    return out_path


# 하위 호환: 기존 호출부(브랜드검색 전용 명칭)도 종합 생성기로 연결
generate_brand_search_report = generate_full_report


# ── 비드앤비 공통 포맷 기반 템플릿 복사·채우기 ──────────────────────

def _w(ws, row: int, col: int, value, alignment=None):
    """기존 셀 서식을 건드리지 않고 value만 교체. 병합 셀은 무시."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    if alignment:
        cell.alignment = alignment


def _save_row_style(ws, row: int, col_range) -> dict:
    """행의 셀 서식(폰트·채우기·테두리·정렬·숫자형식)을 dict로 저장."""
    from openpyxl.cell.cell import MergedCell
    styles = {}
    for col in col_range:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue
        styles[col] = {
            'font':          copy.copy(cell.font),
            'fill':          copy.copy(cell.fill),
            'border':        copy.copy(cell.border),
            'alignment':     copy.copy(cell.alignment),
            'number_format': cell.number_format,
        }
    return styles


def _apply_row_style(ws, row: int, saved_styles: dict):
    """_save_row_style()로 저장한 서식을 다른 행에 그대로 적용."""
    from openpyxl.cell.cell import MergedCell
    for col, style in saved_styles.items():
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue
        cell.font          = style['font']
        cell.fill          = style['fill']
        cell.border        = style['border']
        cell.alignment     = style['alignment']
        cell.number_format = style['number_format']


def _last_styled_row(ws, start_row: int, search_limit: int, check_col: int) -> int:
    """start_row부터 탐색해 테두리가 있는 마지막 행 번호 반환 (없으면 start_row - 1)."""
    from openpyxl.cell.cell import MergedCell
    last = start_row - 1
    for r in range(start_row, start_row + search_limit):
        cell = ws.cell(row=r, column=check_col)
        if isinstance(cell, MergedCell):
            continue
        if cell.border:
            for side in ('left', 'right', 'top', 'bottom'):
                side_obj = getattr(cell.border, side)
                if side_obj and side_obj.style:
                    last = r
                    break
    return last


def generate_using_standard_template(
    report: ClientReport,
    raw_data: dict,
    template_path: str | Path,
    output_dir: str | Path | None = None,
) -> Path:
    """
    비드앤비 공통 포맷(template_path)을 복사한 뒤 실제 데이터로 채워 반환.

    template_path : 2026년_05월_비드앤비 운영보고서.xlsx 같은 표준 양식
    raw_data      : get_monthly_stats() 반환 dict
      {monthly_total, monthly_history, media_breakdown, daily_stats}

    시트 구성:
      표지 · 월별 분석요약 · 네이버 유형별 · N_일별 · N_브랜드검색
    검색광고(파워링크) 시트가 없는 계정은 해당 시트 삭제·비움 처리.
    """
    import warnings
    from datetime import datetime as _dt

    out_dir = Path(output_dir) if output_dir else _OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = (
        f"{report.year}년{report.month:02d}월"
        f"_{report.client_name}_운영보고서.xlsx"
    )
    out_path = out_dir / fname
    shutil.copy2(str(template_path), str(out_path))

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(str(out_path))

    now_dt = _dt.now()
    last_day  = _last_day_of_month(report.year, report.month)
    period    = (
        f"{report.year}-{report.month:02d}-01 ~ "
        f"{report.year}-{report.month:02d}-{last_day:02d}"
    )
    mt        = raw_data.get("monthly_total", {})
    history   = raw_data.get("monthly_history", [])
    campaigns = raw_data.get("media_breakdown", [])
    daily     = raw_data.get("daily_stats", [])
    is_bs_only = mt.get("spend", 0) == 0 and mt.get("impressions", 0) > 0

    # ── 1. 표지 ──────────────────────────────────────────────
    if "표지" in wb.sheetnames:
        wc = wb["표지"]
        _w(wc, 11,  1, f"{report.client_name} {report.month}월 운영데이터 보고서")
        _w(wc, 24, 10, report.homepage or "-")
        _w(wc, 25, 10, "네이버 브랜드검색" if is_bs_only else "네이버 검색광고 + 브랜드검색")
        _w(wc, 26, 10, period)
        _w(wc, 27, 10, report.manager or "-")
        _w(wc, 28, 10, report.email or "-")
        _w(wc, 29, 10, report.phone or "-")

    # ── 2. 월별 분석요약 ─────────────────────────────────────
    # 구조: 헤더 row 23, 데이터 row 24~
    # 컬럼: B=월별 C=노출수 D=클릭수 E=CTR F=CPC G=광고비 H=순위
    #        I=전환수 J=전환율 K=전환당비용 L=전환매출 M=ROAS N=코멘트
    MONTHLY_SHEET  = "월별 분석요약"
    HEADER_ROW     = 23
    DATA_START_ROW = HEADER_ROW + 1

    if MONTHLY_SHEET in wb.sheetnames:
        wm = wb[MONTHLY_SHEET]
        _w(wm, 2, 20, now_dt)  # T2: 작성일
        _w(wm, 3,  2, "네이버 브랜드검색 데이터" if is_bs_only else "네이버 검색광고 + 브랜드검색 데이터")

        # 기존 데이터 행 지우기 (값만, 서식 유지)
        from openpyxl.cell.cell import MergedCell as _MC
        for r in range(DATA_START_ROW, DATA_START_ROW + 20):
            for c in range(2, 21):          # B ~ T
                ws_cell = wm.cell(row=r, column=c)
                if not isinstance(ws_cell, _MC):
                    ws_cell.value = None

        # 새 데이터 쓰기 (3개월 히스토리)
        for i, h in enumerate(history):
            r = DATA_START_ROW + i
            yr, mo = h.get("year", report.year), h.get("month", report.month)
            label = f"{mo}월" if yr == report.year else f"'{str(yr)[2:]}년 {mo}월"
            ctr   = _pct_ratio(h.get("ctr", 0))
            cr    = _pct_ratio(h.get("conversion_rate", 0))

            _w(wm, r,  2, label)
            _w(wm, r,  3, h.get("impressions", 0))
            _w(wm, r,  4, h.get("clicks", 0))
            _w(wm, r,  5, ctr)
            _w(wm, r,  6, h.get("cpc", 0))
            _w(wm, r,  7, h.get("spend", 0))
            _w(wm, r,  8, h.get("avg_rank", 1.0))
            _w(wm, r,  9, h.get("conversions", 0))
            _w(wm, r, 10, cr)
            _w(wm, r, 11, h.get("cpa", 0))
            _w(wm, r, 12, h.get("revenue", 0))
            _w(wm, r, 13, h.get("roas", 0))

        # 코멘트는 N24:R36 병합 영역의 시작 셀(N24)에만 기입
        # 각 월 행(row 25, 26 등)은 MergedCell이라 _w()가 무시하므로 항상 DATA_START_ROW에 씀
        if report.comment:
            comment_cell = wm.cell(row=DATA_START_ROW, column=14)
            comment_cell.value = report.comment
            comment_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── 3. 네이버 유형별 ──────────────────────────────────────
    # 구조: 파워링크 row 6 헤더 / 7 pc / 8 mo
    #        기타   row 16~17 헤더 / 18 파워링크합산 / 19 브랜드검색
    # 컬럼: B=상품 C=노출 D=클릭 E=CTR F=CPC G=광고비 H=순위
    #        I=전환수 J=전환율 K=전환매출액 L=전환당비용
    TYPE_SHEET = "네이버 유형별"
    if TYPE_SHEET in wb.sheetnames:
        wt = wb[TYPE_SHEET]
        _w(wt, 2, 14, now_dt)  # N2: 작성일

        # media_breakdown에서 파워링크/브랜드검색 분리 (태그 기준)
        def _sum_media(lst):
            imp = clk = sp = cv = rv = rw = 0
            for m in lst:
                i = m.get("impressions", 0) or 0
                imp += i; clk += m.get("clicks", 0) or 0
                sp += m.get("spend", 0) or 0; cv += m.get("conversions", 0) or 0
                rv += m.get("revenue", 0) or 0
                rw += (m.get("avg_rank", 0) or 0) * i
            return imp, clk, sp, cv, rv, (rw / imp if imp else 0)

        pl_all = [m for m in campaigns if m.get("media_name", "").startswith("[파워링크]")]
        pl_pc  = [m for m in pl_all if "_PC" in m.get("media_name", "")]
        pl_mo  = [m for m in pl_all if "_MO" in m.get("media_name", "")]
        bs_all = [m for m in campaigns if m.get("media_name", "").startswith("[브랜드검색]")]

        # 단일 계정이거나 태그 없는 경우 fallback
        if not pl_all and not bs_all:
            pl_pc  = [m for m in campaigns if "_PC" in m.get("media_name", "")]
            pl_mo  = [m for m in campaigns if "_MO" in m.get("media_name", "")]
            pl_all = campaigns

        pc_i, pc_c, pc_s, pc_cv, pc_rv, pc_rk = _sum_media(pl_pc)
        mo_i, mo_c, mo_s, mo_cv, mo_rv, mo_rk = _sum_media(pl_mo)
        pl_i, pl_c, pl_s, pl_cv, pl_rv, pl_rk = _sum_media(pl_all)
        bs_i, bs_c, bs_s, bs_cv, bs_rv, bs_rk = _sum_media(bs_all)
        total_cr = _pct_ratio(mt.get("conversion_rate", 0))

        def _write_row(row, imp, clk, sp, cv, rv, rk=None):
            _w(wt, row,  3, imp)
            _w(wt, row,  4, clk)
            _w(wt, row,  5, clk / imp if imp else 0)
            _w(wt, row,  6, round(sp / clk) if clk else 0)
            _w(wt, row,  7, sp)
            if rk is not None:
                _w(wt, row, 8, round(rk, 1))
            _w(wt, row,  9, cv)
            _w(wt, row, 10, cv / clk if clk else 0)
            _w(wt, row, 11, rv)
            _w(wt, row, 12, round(sp / cv) if cv else 0)

        if is_bs_only:
            # 파워링크 없음 → 0
            for r in (7, 8, 18):
                for c in range(3, 13):
                    wt.cell(row=r, column=c).value = 0
        else:
            # row 7: 파워링크 PC
            _write_row(7, pc_i, pc_c, pc_s, pc_cv, pc_rv, pc_rk)
            # row 8: 파워링크 MO
            _write_row(8, mo_i, mo_c, mo_s, mo_cv, mo_rv, mo_rk)
            # row 18: 파워링크 합산 (기타 유형)
            _write_row(18, pl_i, pl_c, pl_s, pl_cv, pl_rv, pl_rk)

        # 브랜드검색 행 (row 19)
        if bs_all:
            _write_row(19, bs_i, bs_c, bs_s, bs_cv, bs_rv, 1.0)
            _w(wt, 19,  6, 0)   # CPC = 0 (고정비)
            _w(wt, 19,  7, 0)   # 광고비 = 0
            _w(wt, 19, 12, 0)   # 전환당비용 = 0
        else:
            # 브랜드검색 태그 없으면 monthly_total 전체를 BS 행에 (단일 BS 계정)
            _w(wt, 19,  3, mt.get("impressions", 0))
            _w(wt, 19,  4, mt.get("clicks", 0))
            _w(wt, 19,  5, mt.get("clicks", 0) / mt.get("impressions", 1) if mt.get("impressions") else 0)
            _w(wt, 19,  6, 0)
            _w(wt, 19,  7, 0)
            _w(wt, 19,  8, 1.0)
            _w(wt, 19,  9, mt.get("conversions", 0))
            _w(wt, 19, 10, total_cr)
            _w(wt, 19, 11, mt.get("revenue", 0))
            _w(wt, 19, 12, 0)

    # ── 4. N_브랜드검색 (단일 시트) or N_브랜드검색_PC / N_브랜드검색_MO ─────────
    # media_breakdown에서 브랜드검색 캠페인 분리
    bs_campaigns = [m for m in campaigns if m.get("media_name", "").startswith("[브랜드검색]")]
    if not bs_campaigns:
        bs_campaigns = campaigns  # 단일 BS 계정 fallback

    from openpyxl.cell.cell import MergedCell as _MC2

    def _fill_bs_detail_sheet(ws_bs, data_rows):
        """브랜드검색 상세 시트 채우기 (단일/PC/MO 공통)."""
        from copy import copy as _cs_bs

        # row 8 서식 참조 저장 (값 초기화 전에 캡처)
        _bs_style = {}
        for c in range(2, 14):
            ref = ws_bs.cell(8, c)
            if not isinstance(ref, _MC2):
                _bs_style[c] = (
                    _cs_bs(ref.font), _cs_bs(ref.fill),
                    _cs_bs(ref.border), _cs_bs(ref.alignment),
                    ref.number_format,
                )

        _w(ws_bs, 2, 14, now_dt)
        # 기존 데이터 행 값만 초기화 (서식 유지)
        max_r = ws_bs.max_row
        for r in range(7, max_r + 1):
            for c in range(2, 14):
                _c = ws_bs.cell(row=r, column=c)
                if not isinstance(_c, _MC2):
                    _c.value = None

        if not data_rows:
            return

        # 합계 행 (row 7)
        tot_i  = sum(m.get("impressions", 0) or 0 for m in data_rows)
        tot_c  = sum(m.get("clicks", 0) or 0 for m in data_rows)
        tot_cv = sum(m.get("conversions", 0) or 0 for m in data_rows)
        tot_rv = sum(m.get("revenue", 0) or 0 for m in data_rows)
        ws_bs.cell(row=7, column=3).value = tot_i
        ws_bs.cell(row=7, column=4).value = tot_c
        r7c5 = ws_bs.cell(7, 5); r7c5.value = tot_c / tot_i if tot_i else 0; r7c5.number_format = "0.00%"
        ws_bs.cell(row=7, column=6).value = 0
        ws_bs.cell(row=7, column=7).value = 0
        ws_bs.cell(row=7, column=8).value = 1.0
        ws_bs.cell(row=7, column=9).value = tot_cv
        r7c10 = ws_bs.cell(7, 10); r7c10.value = tot_cv / tot_c if tot_c else 0; r7c10.number_format = "0.00%"
        ws_bs.cell(row=7, column=11).value = tot_rv
        ws_bs.cell(row=7, column=12).value = 0

        for i, camp in enumerate(data_rows):
            r = 8 + i
            # 템플릿 범위 초과 행에 서식 복사
            for c, s in _bs_style.items():
                cell = ws_bs.cell(r, c)
                if not isinstance(cell, _MC2):
                    cell.font = _cs_bs(s[0]); cell.fill = _cs_bs(s[1])
                    cell.border = _cs_bs(s[2]); cell.alignment = _cs_bs(s[3])
                    cell.number_format = s[4]
            imp = camp.get("impressions", 0) or 0
            clk = camp.get("clicks", 0) or 0
            cv  = camp.get("conversions", 0) or 0
            rv  = camp.get("revenue", 0) or 0
            _w(ws_bs, r,  2, camp.get("media_name", ""))
            _w(ws_bs, r,  3, imp)
            _w(ws_bs, r,  4, clk)
            c5 = ws_bs.cell(r, 5); c5.value = clk / imp if imp else 0; c5.number_format = "0.00%"
            _w(ws_bs, r,  6, 0)
            _w(ws_bs, r,  7, 0)
            _w(ws_bs, r,  8, 1.0)
            _w(ws_bs, r,  9, cv)
            c10 = ws_bs.cell(r, 10); c10.value = cv / clk if clk else 0; c10.number_format = "0.00%"
            _w(ws_bs, r, 11, rv)
            _w(ws_bs, r, 12, 0)
        # 광고그룹명 열 너비 자동 조정
        try:
            from openpyxl.utils import get_column_letter as _gcl2
            col_letter = _gcl2(2)
            max_len = max((len(str(ws_bs.cell(r, 2).value or "")) for r in range(7, ws_bs.max_row+1)), default=10)
            ws_bs.column_dimensions[col_letter].width = min(max(max_len + 2, 15), 50)
        except Exception:
            pass

    # PC/MO 분리 시트 (클라이언트 수정본 템플릿)
    bs_adgroup_pc = raw_data.get("bs_adgroup_pc", [])
    bs_adgroup_mo = raw_data.get("bs_adgroup_mo", [])

    if "N_브랜드검색_PC" in wb.sheetnames:
        _fill_bs_detail_sheet(wb["N_브랜드검색_PC"], bs_adgroup_pc or bs_campaigns)
    if "N_브랜드검색_MO" in wb.sheetnames:
        _fill_bs_detail_sheet(wb["N_브랜드검색_MO"], bs_adgroup_mo or bs_campaigns)

    # 구 템플릿 호환: N_브랜드검색 단일 시트
    if "N_브랜드검색" in wb.sheetnames:
        _fill_bs_detail_sheet(wb["N_브랜드검색"], bs_campaigns)

    # ── 5. N_일별 ────────────────────────────────────────────
    DAILY_SHEET = "N_일별"
    if DAILY_SHEET in wb.sheetnames:
        wd = wb[DAILY_SHEET]
        _w(wd, 2, 15, now_dt)   # O2: 작성일
        if is_bs_only:
            _w(wd, 2, 2, "일별 분석 데이터 (브랜드검색 계정 — 일별 데이터 미제공)")
            # 기존 합계·일별 데이터 지우기
            from openpyxl.cell.cell import MergedCell as _MC3
            for r in range(7, 60):
                for c in range(2, 16):
                    _dc = wd.cell(row=r, column=c)
                    if not isinstance(_dc, _MC3):
                        _dc.value = None
        elif not daily:
            # daily 수집 실패 시 템플릿 잔류 데이터 제거
            from openpyxl.cell.cell import MergedCell as _MC3
            _w(wd, 2, 2, "일별 분석 데이터 (일별 통계 수집 실패 — 재시도 필요)")
            for r in range(7, wd.max_row + 1):
                for c in range(2, 16):
                    _dc = wd.cell(row=r, column=c)
                    if not isinstance(_dc, _MC3):
                        _dc.value = None
        else:
            # 실제 일별 데이터 쓰기
            from openpyxl.cell.cell import MergedCell as _MC3
            for r in range(7, 60):
                for c in range(2, 16):
                    _dc = wd.cell(row=r, column=c)
                    if not isinstance(_dc, _MC3):
                        _dc.value = None
            weekdays = ["월요일","화요일","수요일","목요일","금요일","토요일","일요일"]
            total_row = {"imp": 0, "clk": 0, "spend": 0, "conv": 0, "rev": 0}
            for i, dr in enumerate(daily):
                r = 8 + i
                imp   = dr.get("impressions", 0)
                clk   = dr.get("clicks", 0)
                spend = dr.get("spend", 0)
                conv  = dr.get("conversions", 0)
                rev   = dr.get("revenue", 0)
                d_str = str(dr.get("date", ""))[:10]  # 날짜 부분만 (시간 제거)
                try:
                    from datetime import date as _date
                    from openpyxl.cell.cell import MergedCell as _MCDate
                    d_obj = _date.fromisoformat(d_str)
                    cell = wd.cell(row=r, column=2)
                    if not isinstance(cell, _MCDate):
                        cell.value = d_obj
                        cell.number_format = "YYYY-MM-DD"
                    _w(wd, r, 3, weekdays[d_obj.weekday()])
                except Exception:
                    _w(wd, r, 2, d_str)
                _w(wd, r,  4, imp)
                _w(wd, r,  5, clk)
                _w(wd, r,  6, clk / imp if imp else 0)
                _w(wd, r,  7, spend / clk if clk else 0)
                _w(wd, r,  8, spend)
                _w(wd, r,  9, dr.get("avg_rank", 0))
                _w(wd, r, 10, conv)
                _w(wd, r, 11, conv / clk if clk else 0)
                _w(wd, r, 12, rev)
                _w(wd, r, 13, spend / conv if conv else 0)
                for k, v in (("imp", imp), ("clk", clk), ("spend", spend), ("conv", conv), ("rev", rev)):
                    total_row[k] += v
            # 합계 행 (row 7)
            t = total_row
            _w(wd, 7, 2, "합계")
            _w(wd, 7, 4, t["imp"])
            _w(wd, 7, 5, t["clk"])
            _w(wd, 7, 6, t["clk"] / t["imp"] if t["imp"] else 0)
            _w(wd, 7, 7, t["spend"] / t["clk"] if t["clk"] else 0)
            _w(wd, 7, 8, t["spend"])
            _w(wd, 7, 10, t["conv"])
            _w(wd, 7, 12, t["rev"])

    # ── 6a. N_파워링크PC — 파워링크 전용 키워드 (keyword_stats_pl 우선, fallback keyword_stats)
    from openpyxl.cell.cell import MergedCell as _MCU
    from openpyxl.utils import get_column_letter as _gcl
    PCT_FMT = "0.00%"

    def _auto_col_width(ws, col_num, min_w=10, max_w=50):
        """지정 열의 내용 길이 기준으로 열 너비 자동 조정."""
        col_letter = _gcl(col_num)
        max_len = 0
        for cell in ws[col_letter]:
            if isinstance(cell, _MCU) or cell.value is None:
                continue
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

    keywords = raw_data.get("keyword_stats_pl") or raw_data.get("keyword_stats", [])
    if "N_파워링크PC" in wb.sheetnames:
        wpk = wb["N_파워링크PC"]

        # N_브랜드검색_PC row 8에서 데이터 행 서식 참조 (테두리/배경/폰트)
        from copy import copy as _cs
        _pk_row_style = {}
        _ref_ws = (wb["N_브랜드검색_PC"] if "N_브랜드검색_PC" in wb.sheetnames
                   else wb["N_브랜드검색_MO"] if "N_브랜드검색_MO" in wb.sheetnames
                   else None)
        if _ref_ws:
            for c in range(2, 13):
                _rc = _ref_ws.cell(8, c)
                if not isinstance(_rc, _MCU):
                    _pk_row_style[c] = (
                        _cs(_rc.font), _cs(_rc.fill),
                        _cs(_rc.border), _cs(_rc.alignment),
                        _rc.number_format,
                    )

        def _apply_pk_row_style(ws, row):
            for c, s in _pk_row_style.items():
                cell = ws.cell(row, c)
                if not isinstance(cell, _MCU):
                    cell.font, cell.fill, cell.border, cell.alignment = (
                        _cs(s[0]), _cs(s[1]), _cs(s[2]), _cs(s[3])
                    )
                    cell.number_format = s[4]

        if keywords:
            # 제목 업데이트
            tc = wpk.cell(row=2, column=2)
            if not isinstance(tc, _MCU):
                tc.value = "파워링크 키워드별 분석 데이터 (PC+MO 합산)"
            # 데이터 행 서식+값 초기화 (정확히 keywords 개수만큼만)
            for r in range(8, 8 + len(keywords)):
                _apply_pk_row_style(wpk, r)
                for c in range(2, 13):
                    cell = wpk.cell(r, c)
                    if not isinstance(cell, _MCU):
                        cell.value = None
            # 기존 템플릿 잔여 행 값 정리 (서식은 유지, 값만 삭제)
            for r in range(8 + len(keywords), wpk.max_row + 1):
                for c in range(2, 13):
                    cell = wpk.cell(r, c)
                    if not isinstance(cell, _MCU):
                        cell.value = None
            for i, kw in enumerate(keywords):
                r = 8 + i
                imp   = kw.get("impressions", 0) or 0
                clk   = kw.get("clicks", 0) or 0
                spend = kw.get("spend", 0) or 0
                conv  = kw.get("conversions", 0) or 0
                rev   = kw.get("revenue", 0) or 0
                _w(wpk, r, 2, kw.get("keyword", ""))
                _w(wpk, r, 3, imp)
                _w(wpk, r, 4, clk)
                c5 = wpk.cell(r, 5); c5.value = clk / imp if imp else 0; c5.number_format = PCT_FMT
                _w(wpk, r, 6, round(spend / clk) if clk else 0)
                _w(wpk, r, 7, spend)
                _w(wpk, r, 8, kw.get("avg_rank", 0) or 0)
                _w(wpk, r, 9, conv)
                c10 = wpk.cell(r, 10); c10.value = conv / clk if clk else 0; c10.number_format = PCT_FMT
                _w(wpk, r, 11, rev)
                _w(wpk, r, 12, round(spend / conv) if conv else 0)
            # 합계 행 (row 7)
            tot_imp = sum(k.get("impressions",0) or 0 for k in keywords)
            tot_clk = sum(k.get("clicks",0) or 0 for k in keywords)
            tot_sp  = sum(k.get("spend",0) or 0 for k in keywords)
            tot_cv  = sum(k.get("conversions",0) or 0 for k in keywords)
            tot_rv  = sum(k.get("revenue",0) or 0 for k in keywords)
            _w(wpk, 7, 3, tot_imp); _w(wpk, 7, 4, tot_clk)
            c7_5 = wpk.cell(7, 5); c7_5.value = tot_clk/tot_imp if tot_imp else 0; c7_5.number_format = PCT_FMT
            _w(wpk, 7, 6, round(tot_sp/tot_clk) if tot_clk else 0)
            _w(wpk, 7, 7, tot_sp); _w(wpk, 7, 9, tot_cv)
            c7_10 = wpk.cell(7, 10); c7_10.value = tot_cv/tot_clk if tot_clk else 0; c7_10.number_format = PCT_FMT
            _w(wpk, 7, 11, tot_rv); _w(wpk, 7, 12, round(tot_sp/tot_cv) if tot_cv else 0)
            # 키워드명 열 너비 자동 조정
            _auto_col_width(wpk, 2, min_w=15, max_w=40)
        else:
            tc = wpk.cell(row=2, column=2)
            if not isinstance(tc, _MCU) and tc.value and "(미제공)" not in str(tc.value):
                tc.value = str(tc.value) + "  (미제공)"
            for r in range(8, wpk.max_row + 1):
                for c in range(2, 13):
                    cell = wpk.cell(r, c)
                    if not isinstance(cell, _MCU):
                        cell.value = None
            mc = wpk.cell(row=8, column=2)
            if not isinstance(mc, _MCU):
                mc.value = "데이터 미제공  (파워링크 계정 미연동)"
                mc.font = Font(name="맑은 고딕", size=10, color="888888", italic=True)

    # N_파워링크MO — 키워드별 디바이스 분리 미지원 안내
    if "N_파워링크MO" in wb.sheetnames:
        wmo = wb["N_파워링크MO"]
        tc = wmo.cell(row=2, column=2)
        if not isinstance(tc, _MCU):
            tc.value = "파워링크 키워드별 분석 데이터  (MO 분리 미지원)"
        if wmo.max_row >= 8:
            wmo.delete_rows(8, wmo.max_row - 8 + 1)
        wmo.insert_rows(8)
        mc = wmo.cell(row=8, column=2)
        if not isinstance(mc, _MCU):
            mc.value = "MO 디바이스 분리 통계 미지원 — PC+MO 합산 데이터는 N_파워링크PC 시트 참조"
            mc.font = Font(name="맑은 고딕", size=10, color="888888", italic=True)
            mc.alignment = Alignment(horizontal="left", vertical="center")

    # ── 6b. N_시간대별 (breakdown=hh24, 최근 7일) ──────────────
    hourly_stats = raw_data.get("hourly_stats", [])
    hourly_since = raw_data.get("hourly_since", "")
    hourly_until = raw_data.get("hourly_until", "")

    if "N_시간대별" in wb.sheetnames:
        wh = wb["N_시간대별"]
        from copy import copy as _cs_h
        # row 8 서식 참조 (before clear)
        _h_style = {}
        for c in range(2, 13):
            ref = wh.cell(8, c)
            if not isinstance(ref, _MCU):
                _h_style[c] = (_cs_h(ref.font), _cs_h(ref.fill),
                               _cs_h(ref.border), _cs_h(ref.alignment), ref.number_format)

        # 제목 업데이트
        tc = wh.cell(row=2, column=2)
        if not isinstance(tc, _MCU):
            period = f"  (최근 7일: {hourly_since} ~ {hourly_until})" if hourly_since else ""
            base = str(tc.value or "시간대별 분석").split("  (")[0]
            tc.value = base + period

        # 기존 데이터 행 삭제 후 재삽입
        if wh.max_row >= 7:
            wh.delete_rows(7, wh.max_row - 7 + 1)

        if hourly_stats:
            # 합계 행 (row 7)
            tot_imp = sum(h.get("impressions", 0) or 0 for h in hourly_stats)
            tot_clk = sum(h.get("clicks", 0) or 0 for h in hourly_stats)
            tot_sp  = sum(h.get("spend", 0) or 0 for h in hourly_stats)
            tot_cv  = sum(h.get("conversions", 0) or 0 for h in hourly_stats)
            tot_rv  = sum(h.get("revenue", 0) or 0 for h in hourly_stats)
            # 템플릿 열 구조: col2=시간대, col3=빈칸, col4=노출수, col5=클릭수,
            #   col6=클릭률, col7=클릭당비용, col8=광고비용, col9=평균순위,
            #   col10=전환수, col11=전환율, col12=전환매출액, col13=전환당비용
            wh.insert_rows(7)
            _w(wh, 7, 2, "합계")
            _w(wh, 7, 4, tot_imp); _w(wh, 7, 5, tot_clk)
            c7_6 = wh.cell(7, 6); c7_6.value = tot_clk/tot_imp if tot_imp else 0; c7_6.number_format = PCT_FMT
            _w(wh, 7, 7, round(tot_sp/tot_clk) if tot_clk else 0)
            _w(wh, 7, 8, tot_sp); _w(wh, 7, 10, tot_cv)
            c7_11 = wh.cell(7, 11); c7_11.value = tot_cv/tot_clk if tot_clk else 0; c7_11.number_format = PCT_FMT
            _w(wh, 7, 12, tot_rv); _w(wh, 7, 13, round(tot_sp/tot_cv) if tot_cv else 0)

            # 24개 시간대 데이터 행
            for i, hr in enumerate(hourly_stats):
                r = 8 + i
                wh.insert_rows(r)
                for c, s in _h_style.items():
                    cell = wh.cell(r, c)
                    if not isinstance(cell, _MCU):
                        cell.font = _cs_h(s[0]); cell.fill = _cs_h(s[1])
                        cell.border = _cs_h(s[2]); cell.alignment = _cs_h(s[3])
                        cell.number_format = s[4]
                imp = hr.get("impressions", 0) or 0
                clk = hr.get("clicks", 0) or 0
                sp  = hr.get("spend", 0) or 0
                cv  = hr.get("conversions", 0) or 0
                rv  = hr.get("revenue", 0) or 0
                _w(wh, r, 2, hr.get("hour", ""))
                _w(wh, r, 4, imp); _w(wh, r, 5, clk)
                ch6 = wh.cell(r, 6); ch6.value = clk/imp if imp else 0; ch6.number_format = PCT_FMT
                _w(wh, r, 7, round(sp/clk) if clk else 0)
                _w(wh, r, 8, sp)
                _w(wh, r, 9, hr.get("avg_rank", 0) or 0)
                _w(wh, r, 10, cv)
                ch11 = wh.cell(r, 11); ch11.value = cv/clk if clk else 0; ch11.number_format = PCT_FMT
                _w(wh, r, 12, rv)
                _w(wh, r, 13, round(sp/cv) if cv else 0)
        else:
            wh.insert_rows(7)
            msg = wh.cell(7, 2)
            if not isinstance(msg, _MCU):
                msg.value = "시간대별 데이터 없음 (최근 7일 기준, 데이터 미수집)"
                msg.font = Font(name="맑은 고딕", size=10, color="888888", italic=True)

    # ── TG 광고 미제공 처리 ──────────────────────────────────
    NOT_YET_IMPLEMENTED = {
        "TG 광고": (2, 2, 7),
    }

    for sname, (title_r, title_c, data_start) in NOT_YET_IMPLEMENTED.items():
        if sname not in wb.sheetnames:
            continue
        wu = wb[sname]
        tc = wu.cell(row=title_r, column=title_c)
        if not isinstance(tc, _MCU) and tc.value and "(미제공)" not in str(tc.value):
            tc.value = str(tc.value) + "  (미제공)"
        if wu.max_row >= data_start:
            wu.delete_rows(data_start, wu.max_row - data_start + 1)
        wu.insert_rows(data_start)
        msg_cell = wu.cell(row=data_start, column=title_c)
        if not isinstance(msg_cell, _MCU):
            msg_cell.value = "데이터 미제공  (해당 계정에서 제공하지 않는 매체입니다)"
            msg_cell.font = Font(name="맑은 고딕", size=10, color="888888", italic=True)
            msg_cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(str(out_path))
    return out_path
