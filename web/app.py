"""
광고 보고서B 자동화 웹 앱 (정리본)
핵심 동선만: 로그인 → 광고주 등록(업로드/개별/검색/삭제) → 버전B 생성·다운로드
계정: /setup(최초 admin) · /admin/users(팀원 관리)
"""
import asyncio
import hashlib
import os
import sys
import uuid
import warnings
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

from fastapi import FastAPI, Request, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from itsdangerous import URLSafeTimedSerializer, BadSignature
from dotenv import load_dotenv

ROOT = Path(__file__).parent.parent
SRC  = ROOT / "src"
sys.path.insert(0, str(SRC))

load_dotenv(ROOT / ".env")

from web.db import (
    init_db, get_all_clients, get_clients_by_owner, get_client,
    upsert_client, delete_client, delete_clients_by_owner,
    create_report, update_report, get_report, list_reports, delete_report,
    enqueue_report, claim_next_report, requeue_stuck, report_status_counts,
    set_key_status, last_report_dates, get_user_names,
    get_team_ids, get_members_of, get_users_by_role, get_clients_scoped,
    set_parent, assign_report,
)
from collectors.naver_searchad import is_connected, MOCK_MODE

init_db()

app = FastAPI(title="보고서B 자동화")
# 정적 폴더가 없으면(예: git 빈 폴더 미커밋) 생성 후 마운트 — 없으면 시작 시 크래시함
_STATIC_DIR = Path(__file__).parent / "static"
_STATIC_DIR.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(_STATIC_DIR)), name="static")
templates = Jinja2Templates(directory=str(Path(__file__).parent / "templates"))

# DATA_DIR: 영속 저장 루트. Fly 볼륨(/app/data) 마운트 시 여기에 DB·보고서·업로드가 모두 남는다.
# 로컬은 미설정 → 프로젝트 루트 사용(기존 동작 유지).
DATA_DIR = Path(os.getenv("DATA_DIR", str(ROOT)))
OUTPUT_DIR = DATA_DIR / "output"
UPLOAD_DIR = DATA_DIR / "uploads"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

SECRET_KEY = os.getenv("SESSION_SECRET", "change-me-in-production")
signer = URLSafeTimedSerializer(SECRET_KEY)


def render(request: Request, name: str, **ctx):
    """로그인 사용자용 페이지 렌더 — 공통 컨텍스트 자동 주입."""
    sess = get_session(request)
    now = datetime.now()
    base_ctx = {
        "sess": sess,
        "api_connected": is_connected(),
        "mock_mode": MOCK_MODE,
        "now_year": now.year,
        "now_month": now.month,
    }
    base_ctx.update(ctx)
    return templates.TemplateResponse(request=request, name=name, context=base_ctx)


# 진행 중 작업 (메모리)
jobs: dict[str, dict] = {}


# ── 세션 / 인증 ─────────────────────────────────────────
# 비밀번호 해싱용 고정 솔트 — SESSION_SECRET 과 분리한다.
# (세션키를 바꿔도 비밀번호/로그인이 깨지지 않게 하기 위함)
PW_SALT = os.getenv("PASSWORD_SALT", "e-noble-report::pw-salt::v1")


def _hash_pw(pw: str) -> str:
    return hashlib.pbkdf2_hmac("sha256", pw.encode(), PW_SALT.encode(), 100_000).hex()


def get_session(request: Request) -> dict | None:
    token = request.cookies.get("session")
    if not token:
        return None
    try:
        return signer.loads(token, max_age=86400 * 7)
    except BadSignature:
        return None


def require_session(request: Request) -> dict:
    sess = get_session(request)
    if not sess:
        raise HTTPException(status_code=302, headers={"Location": "/login"})
    return sess


def _set_session_cookie(response, user: dict):
    data = {"user_id": user["id"], "name": user["name"], "role": user["role"]}
    response.set_cookie("session", signer.dumps(data), httponly=True, max_age=86400 * 7)


def _is_admin(sess) -> bool:
    return bool(sess and sess.get("role") == "admin")


def _visibility(sess):
    """(owner_ids, member_self) 반환.
    - admin  → (None, None)         전체
    - manager→ (팀 id들, None)       자기 + 소속 팀원 소유
    - member → ([자기], 자기)        자기 소유 + 나에게 할당된 보고서
    """
    role = sess.get("role")
    uid = sess.get("user_id")
    if role == "admin":
        return None, None
    if role == "manager":
        return get_team_ids(uid), None
    return [uid], uid


def _owns(sess, owner_id) -> bool:
    """광고주 편집/생성/삭제 권한: admin=전체, 그 외=가시 범위(팀/본인) 내."""
    if _is_admin(sess):
        return True
    owner_ids, _ = _visibility(sess)
    return owner_id is not None and owner_id in (owner_ids or [])


def _can_see_report(sess, r: dict) -> bool:
    """보고서 열람 권한(다운로드/삭제/재시도용)."""
    if _is_admin(sess):
        return True
    owner_ids, member_self = _visibility(sess)
    if r.get("owner_id") in (owner_ids or []):
        return True
    return member_self is not None and r.get("assigned_to") == member_self


# ── 로그인 ──────────────────────────────────────────────
@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    if get_session(request):
        return RedirectResponse("/", status_code=302)
    return templates.TemplateResponse(request=request, name="login.html", context={"error": None})


@app.post("/login")
async def login(request: Request, email: str = Form(...), password: str = Form(...)):
    from web.db import get_conn
    with get_conn() as conn:
        user = conn.execute(
            "SELECT * FROM users WHERE email=? AND password_hash=?",
            (email.strip().lower(), _hash_pw(password))
        ).fetchone()
    if not user:
        return templates.TemplateResponse(request=request, name="login.html",
                                          context={"error": "이메일 또는 비밀번호가 틀렸습니다."})
    resp = RedirectResponse("/", status_code=302)
    _set_session_cookie(resp, dict(user))
    return resp


@app.get("/logout")
async def logout():
    resp = RedirectResponse("/login", status_code=302)
    resp.delete_cookie("session")
    return resp


# ── 루트: 로그인 후 광고주 페이지로 ──────────────────────
@app.get("/")
async def index(request: Request):
    if not get_session(request):
        return RedirectResponse("/login", status_code=302)
    return RedirectResponse("/clients", status_code=302)


# ── 광고주 등록·검색·삭제 ─────────────────────────────────
async def _bg_validate_keys(items: list):
    """업로드된 키들을 순차 검증(조회 전용 1회)해 key_status 갱신."""
    from collectors.naver_searchad import validate_account
    for (cid_db, cid, key, secret) in items:
        try:
            ok = await asyncio.to_thread(validate_account, cid, key, secret)
            set_key_status(cid_db, "ok" if ok else "invalid")
        except Exception:
            pass


def _mask(secret) -> str:
    s = str(secret or "")
    if not s:
        return ""
    return "••••" + s[-4:] if len(s) > 4 else "••••"


@app.get("/clients", response_class=HTMLResponse)
async def clients_page(request: Request):
    sess = get_session(request)
    if not sess:
        return RedirectResponse("/login", status_code=302)
    ov, _ms = _visibility(sess)
    clients = get_all_clients() if ov is None else get_clients_scoped(ov)
    last_dates = last_report_dates(ov)
    user_names = get_user_names() if sess["role"] in ("admin", "manager") else {}
    rows = []
    for c in clients:
        rows.append({
            "id": c["id"], "name": c["name"],
            "naver_customer_id": c.get("naver_customer_id") or "",
            "media": c.get("media") or "",
            "manager_name": c.get("manager_name") or "",
            "has_key": bool(c.get("api_key") and c.get("api_secret")),
            "key_masked": _mask(c.get("api_key")),
            "key_status": c.get("key_status") or "",
            "last_report": last_dates.get(c["id"], ""),
            "owner_name": user_names.get(c.get("owner_id"), "-"),
        })
    return render(request, "clients.html", active_page="clients",
                  client_rows=rows,
                  added=request.query_params.get("added"),
                  uploaded=request.query_params.get("uploaded"),
                  err=request.query_params.get("err"))


@app.post("/clients/add")
async def clients_add(
    request: Request,
    name: str = Form(...),
    naver_customer_id: str = Form(...),
    api_key: str = Form(""),
    api_secret: str = Form(""),
    media: str = Form("파워링크"),
    manager_name: str = Form(""),
    manager_email: str = Form(""),
):
    sess = get_session(request)
    if not sess:
        return RedirectResponse("/login", status_code=302)
    cid_db, _ = upsert_client(
        naver_customer_id=naver_customer_id.strip(), name=name.strip(),
        owner_id=sess["user_id"],
        api_key=api_key.strip(), api_secret=api_secret.strip(),
        media=media.strip(), manager_name=manager_name.strip(),
        manager_email=manager_email.strip(),
    )
    # 키 유효성 검증(조회 전용 1회) → 대시보드 '정보 없음' 표기용
    if api_key.strip() and api_secret.strip():
        from collectors.naver_searchad import validate_account
        ok = await asyncio.to_thread(validate_account, naver_customer_id.strip(),
                                     api_key.strip(), api_secret.strip())
        set_key_status(cid_db, "ok" if ok else "invalid")
    return RedirectResponse(f"/clients?added={name.strip()}", status_code=302)


@app.post("/clients/{client_id}/delete")
async def clients_delete(client_id: int, request: Request):
    sess = get_session(request)
    if not sess:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    c = get_client(client_id)
    if not c:
        return JSONResponse({"error": "없는 광고주"}, status_code=404)
    if not _owns(sess, c.get("owner_id")):
        return JSONResponse({"error": "권한 없음"}, status_code=403)
    delete_client(client_id)
    return JSONResponse({"ok": True})


@app.get("/clients/template")
async def clients_template(request: Request):
    """로스터 업로드용 xlsx 템플릿 다운로드."""
    if not get_session(request):
        return RedirectResponse("/login", status_code=302)
    import roster
    tpl = UPLOAD_DIR / "roster_template.xlsx"
    roster.create_template(tpl, overwrite=True)
    return FileResponse(path=str(tpl), filename="광고주_로스터_템플릿.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/clients/upload")
async def clients_upload(request: Request, file: UploadFile = File(...),
                         mode: str = Form("append")):
    """엑셀/CSV 로스터 업로드 → 광고주 일괄 등록.
    mode='append' : 기존 목록에 없는 광고주 추가(있으면 갱신)
    mode='replace': 내 광고주 전체 삭제 후 시트 데이터로 새로 채움
    """
    sess = get_session(request)
    if not sess:
        return RedirectResponse("/login", status_code=302)
    import roster
    dest = UPLOAD_DIR / f"roster_{uuid.uuid4().hex[:8]}_{file.filename}"
    dest.write_bytes(await file.read())
    try:
        records = roster.read_file(dest)
    except Exception as e:
        return RedirectResponse(f"/clients?err={e}", status_code=302)

    valid = [r for r in records if (r.get("customer_id") or "").strip()]
    if mode == "replace" and not valid:
        return RedirectResponse("/clients?err=시트에 유효한 광고주(customer_id)가 없어 전체 교체를 중단했습니다.",
                                status_code=302)

    deleted = delete_clients_by_owner(sess["user_id"]) if mode == "replace" else 0

    new_cnt, upd_cnt, skipped = 0, 0, len(records) - len(valid)
    to_validate = []
    for r in valid:
        cid = r["customer_id"].strip()
        k, s = r.get("api_key", "").strip(), r.get("api_secret", "").strip()
        _id, created = upsert_client(
            naver_customer_id=cid, name=r.get("name") or f"광고주 {cid}",
            owner_id=sess["user_id"],
            api_key=k, api_secret=s,
            media=r.get("media", ""), manager_name=r.get("marketer", ""),
            manager_email=r.get("email", ""),
        )
        if created:
            new_cnt += 1
        else:
            upd_cnt += 1
        if k and s:
            to_validate.append((_id, cid, k, s))

    # 업로드된 키는 백그라운드로 검증(응답 지연 방지) → 대시보드에 '정보 없음' 반영
    if to_validate:
        asyncio.create_task(_bg_validate_keys(to_validate))

    if mode == "replace":
        msg = f"기존 {deleted}건 삭제 · {new_cnt}건 추가"
    else:
        msg = f"{new_cnt}건 추가" + (f" · {upd_cnt}건 갱신" if upd_cnt else "")
    if skipped:
        msg += f" · {skipped}건 건너뜀"
    return RedirectResponse(f"/clients?uploaded={msg}", status_code=302)


# ── 버전B 보고서 생성 (서버측 큐) ────────────────────────
# 팀장이 '생성/전체 생성'을 누르면 즉시 큐에 등록만 하고 끝난다.
# 실제 생성은 서버 백그라운드 워커가 처리하므로, 페이지를 나가거나
# 로그아웃해도 전부 만들어져 [생성된 보고서]에 나타난다.
QUEUE_CONCURRENCY = int(os.getenv("QUEUE_CONCURRENCY", "2"))
_running_tasks: set = set()


def _read_summary_comment(filepath) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath))
    if "파워링크_Summary" in wb.sheetnames:
        return wb["파워링크_Summary"].cell(row=9, column=12).value or ""
    return ""


async def _process_report(report: dict):
    """큐에서 잡은 보고서 1건 생성. 실패는 status='error'+사유 로 남긴다."""
    rid = report["id"]
    client = get_client(report["client_id"])
    if not client:
        update_report(rid, status="error", error="광고주가 삭제되어 생성할 수 없습니다.")
        return
    year, month = report["year"], report["month"]
    try:
        if client.get("api_key") and client.get("api_secret"):
            from collectors.naver_searchad import register_account
            register_account(
                client.get("naver_customer_id"),
                client["api_key"], client["api_secret"],
                name=client.get("name", ""), media=client.get("media", ""),
            )
        from v2_report import generate_v2_report
        uniq = f"{year}년{month:02d}월_{client['name']}_보고서B_{rid}.xlsx"
        out_path = await asyncio.to_thread(
            generate_v2_report,
            client.get("naver_customer_id") or str(client["id"]),
            year, month,
            client_name=client["name"], output_dir=OUTPUT_DIR, out_name=uniq,
        )
        filename = out_path.name
        comment = await asyncio.to_thread(_read_summary_comment, OUTPUT_DIR / filename)
        update_report(rid, status="done", filename=filename, comment=comment or "")
    except Exception as e:
        update_report(rid, status="error", error=str(e)[:300])


async def _worker_loop():
    """대기열 폴링 → 여유만큼 동시 처리(QUEUE_CONCURRENCY)."""
    while True:
        try:
            while len(_running_tasks) < QUEUE_CONCURRENCY:
                rep = claim_next_report()      # queued -> processing (원자적)
                if not rep:
                    break
                t = asyncio.create_task(_process_report(rep))
                _running_tasks.add(t)
                t.add_done_callback(_running_tasks.discard)
        except Exception:
            pass
        await asyncio.sleep(2)


@app.on_event("startup")
async def _start_worker():
    # 크래시/재시작으로 끊긴 작업을 다시 대기열로 되돌리고 워커 기동
    try:
        requeue_stuck()
    except Exception:
        pass
    asyncio.create_task(_worker_loop())


@app.post("/api/generate-v2")
async def generate_v2(
    request: Request,
    client_id: int = Form(...),
    year: int = Form(...),
    month: int = Form(...),
):
    """개별 광고주 1건을 큐에 등록."""
    sess = require_session(request)
    client = get_client(client_id)
    if not client:
        return JSONResponse({"error": "클라이언트 없음"}, status_code=404)
    if not _owns(sess, client.get("owner_id")):
        return JSONResponse({"error": "권한 없음"}, status_code=403)
    if not (client.get("api_key") and client.get("api_secret")):
        return JSONResponse({"error": "API 키가 없어 생성할 수 없습니다. 키를 먼저 등록하세요."},
                            status_code=400)
    rid = enqueue_report(client_id, year, month)
    return JSONResponse({"report_id": rid, "queued": True})


@app.post("/api/generate-batch")
async def generate_batch(request: Request, year: int = Form(...), month: int = Form(...)):
    """내 광고주(키 등록된 것) 전부를 한 번에 큐 등록."""
    sess = require_session(request)
    clients = get_all_clients() if _is_admin(sess) else get_clients_by_owner(sess["user_id"])
    keyed = [c for c in clients if c.get("api_key") and c.get("api_secret")]
    for c in keyed:
        enqueue_report(c["id"], year, month)
    return JSONResponse({"queued": len(keyed), "skipped": len(clients) - len(keyed)})


@app.get("/api/queue-status")
async def queue_status(request: Request):
    """현재 사용자 기준 상태별 개수(예약/진행/완료/실패)."""
    sess = get_session(request)
    if not sess:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    ov, ms = _visibility(sess)
    return JSONResponse(report_status_counts(ov, ms))


# ── 생성된 보고서 ─────────────────────────────────────────
@app.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request):
    sess = get_session(request)
    if not sess:
        return RedirectResponse("/login", status_code=302)
    ov, ms = _visibility(sess)
    reports = list_reports(ov, ms)
    un = get_user_names()
    show_owner = sess["role"] in ("admin", "manager")
    for r in reports:
        if show_owner:
            r["owner_name"] = un.get(r.get("owner_id"), "-")
        r["assignee_name"] = un.get(r.get("assigned_to")) if r.get("assigned_to") else ""
    c = report_status_counts(ov, ms)
    # 할당 대상(팀원) 목록 — 팀장은 소속 팀원, 관리자는 전체 팀원
    if _is_admin(sess):
        team_members = get_users_by_role("member")
    elif sess["role"] == "manager":
        team_members = get_members_of(sess["user_id"])
    else:
        team_members = []
    return render(request, "reports.html", active_page="reports",
                  reports=reports, done_count=c["done"], failed_count=c["error"],
                  queued_count=c["queued"], processing_count=c["processing"],
                  active_count=c["active"], total_count=len(reports),
                  can_assign=bool(team_members), team_members=team_members)


@app.post("/reports/download-zip")
async def reports_download_zip(request: Request, ids: str = Form("")):
    """선택(또는 전체) 완료 보고서를 zip 으로 묶어 다운로드."""
    sess = get_session(request)
    if not sess:
        return RedirectResponse("/login", status_code=302)
    import zipfile
    want = {i for i in ids.split(",") if i.strip()}
    ov, ms = _visibility(sess)
    reports = list_reports(ov, ms)
    picked = [r for r in reports
              if r["status"] == "done" and r["filename"]
              and (not want or str(r["id"]) in want)]
    if not picked:
        return JSONResponse({"error": "다운로드할 완료 보고서가 없습니다."}, status_code=404)

    zpath = UPLOAD_DIR / f"reports_{uuid.uuid4().hex[:8]}.zip"
    with zipfile.ZipFile(zpath, "w", zipfile.ZIP_DEFLATED) as z:
        for r in picked:
            fp = OUTPUT_DIR / r["filename"]
            if fp.exists():
                z.write(fp, arcname=r["filename"])   # 파일명이 report_id 로 유일 → 충돌 없음
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    return FileResponse(path=str(zpath), filename=f"보고서_{stamp}.zip",
                        media_type="application/zip")


@app.post("/reports/{report_id}/retry")
async def report_retry(report_id: int, request: Request):
    """실패(또는 개별) 보고서를 큐에 다시 넣는다. 연·월 그대로 재사용."""
    sess = get_session(request)
    if not sess:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    r = get_report(report_id)
    if not r or not _owns(sess, r.get("owner_id")):
        return JSONResponse({"error": "권한 없음"}, status_code=403)
    update_report(report_id, status="queued", error="")
    return JSONResponse({"ok": True})


@app.post("/reports/retry-failed")
async def reports_retry_failed(request: Request):
    """내 실패 보고서 전부를 다시 큐에 넣는다."""
    sess = get_session(request)
    if not sess:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    n = 0
    ov, ms = _visibility(sess)
    for r in list_reports(ov, ms):
        if r["status"] == "error" and _owns(sess, r.get("owner_id")):
            update_report(r["id"], status="queued", error="")
            n += 1
    return JSONResponse({"queued": n})


@app.post("/reports/delete")
async def reports_delete(request: Request, ids: str = Form("")):
    """보고서 삭제(개별/선택). 소유자(또는 admin)만. 파일도 함께 제거."""
    sess = get_session(request)
    if not sess:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    targets = [i for i in ids.split(",") if i.strip()]
    deleted = 0
    for rid in targets:
        try:
            r = get_report(int(rid))
        except ValueError:
            continue
        if not r or not _owns(sess, r.get("owner_id")):
            continue
        if r.get("filename"):
            fp = OUTPUT_DIR / r["filename"]
            try:
                if fp.exists():
                    fp.unlink()
            except OSError:
                pass
        delete_report(int(rid))
        deleted += 1
    return JSONResponse({"ok": True, "deleted": deleted})


@app.post("/reports/assign")
async def reports_assign(request: Request, ids: str = Form(""), member_id: int = Form(...)):
    """선택한 보고서를 팀원에게 할당(열람 권한 부여). 팀장/관리자만."""
    sess = get_session(request)
    if not sess:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    if sess["role"] not in ("admin", "manager"):
        return JSONResponse({"error": "권한 없음"}, status_code=403)
    if _is_admin(sess):
        valid = {m["id"] for m in get_users_by_role("member")}
    else:
        valid = {m["id"] for m in get_members_of(sess["user_id"])}
    if member_id not in valid:
        return JSONResponse({"error": "내 팀원이 아닙니다."}, status_code=403)
    n = 0
    for rid in ids.split(","):
        rid = rid.strip()
        if not rid:
            continue
        try:
            r = get_report(int(rid))
        except ValueError:
            continue
        if r and _owns(sess, r.get("owner_id")):   # 내가 관리하는 보고서만 할당
            assign_report(int(rid), member_id)
            n += 1
    return JSONResponse({"assigned": n})


# ── 다운로드 ─────────────────────────────────────────────
@app.get("/download/{filename}")
async def download(filename: str, request: Request):
    sess = get_session(request)
    if not sess:
        return RedirectResponse("/login", status_code=302)
    # 볼 수 있는 보고서(본인/팀/할당) 파일만 허용
    ov, ms = _visibility(sess)
    allowed = {r["filename"] for r in list_reports(ov, ms) if r.get("filename")}
    if filename not in allowed:
        raise HTTPException(status_code=404)
    path = OUTPUT_DIR / filename
    if not path.exists():
        raise HTTPException(status_code=404)
    return FileResponse(path=str(path), filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── 사용 매뉴얼 ───────────────────────────────────────────
@app.get("/manual", response_class=HTMLResponse)
async def manual_page(request: Request):
    if not get_session(request):
        return RedirectResponse("/login", status_code=302)
    return render(request, "manual.html", active_page="manual")


# ── 팀원(로그인 계정) 관리 ────────────────────────────────
@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users(request: Request):
    sess = get_session(request)
    if not sess or sess["role"] != "admin":
        return RedirectResponse("/", status_code=302)
    from web.db import get_conn
    with get_conn() as conn:
        users = [dict(r) for r in conn.execute(
            "SELECT * FROM users ORDER BY created_at DESC").fetchall()]
    return render(request, "admin_users.html", active_page="users",
                  users=users, managers=get_users_by_role("manager"),
                  error=None, success=None)


@app.post("/admin/users/create")
async def create_user(
    request: Request,
    name: str = Form(...), email: str = Form(...),
    password: str = Form(...), role: str = Form(...),
):
    sess = get_session(request)
    if not sess or sess["role"] != "admin":
        return RedirectResponse("/", status_code=302)
    from web.db import get_conn
    email = email.strip().lower()
    role = role if role in ("admin", "manager", "member") else "member"
    with get_conn() as conn:
        if conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone():
            users = [dict(r) for r in conn.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()]
            return render(request, "admin_users.html", active_page="users",
                          users=users, managers=get_users_by_role("manager"),
                          error=f"이미 존재하는 이메일입니다: {email}", success=None)
        conn.execute("INSERT INTO users (email, password_hash, name, role) VALUES (?,?,?,?)",
                     (email, _hash_pw(password), name.strip(), role))
    with get_conn() as conn:
        users = [dict(r) for r in conn.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()]
    return render(request, "admin_users.html", active_page="users",
                  users=users, managers=get_users_by_role("manager"),
                  error=None, success=f"{name} 계정이 생성되었습니다.")


@app.post("/admin/users/{user_id}/delete")
async def delete_user(user_id: int, request: Request):
    sess = get_session(request)
    if not sess or sess["role"] != "admin":
        return JSONResponse({"error": "권한 없음"}, status_code=403)
    if user_id == sess["user_id"]:
        return JSONResponse({"error": "본인 계정은 삭제 불가"}, status_code=400)
    from web.db import get_conn
    with get_conn() as conn:
        conn.execute("DELETE FROM users WHERE id=?", (user_id,))
    return JSONResponse({"ok": True})


@app.post("/admin/users/{user_id}/reset-password")
async def reset_user_password(user_id: int, request: Request, password: str = Form(...)):
    """관리자가 팀원 비밀번호를 재설정(계정·소속 광고주 유지)."""
    sess = get_session(request)
    if not sess or sess["role"] != "admin":
        return JSONResponse({"error": "권한 없음"}, status_code=403)
    if len(password.strip()) < 4:
        return JSONResponse({"error": "비밀번호는 4자 이상"}, status_code=400)
    from web.db import get_conn
    with get_conn() as conn:
        conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                     (_hash_pw(password.strip()), user_id))
    return JSONResponse({"ok": True})


@app.post("/admin/users/{user_id}/set-parent")
async def set_user_parent(user_id: int, request: Request, parent_id: str = Form("")):
    """관리자: 팀원의 소속 팀장 지정(빈 값이면 해제)."""
    sess = get_session(request)
    if not sess or sess["role"] != "admin":
        return JSONResponse({"error": "권한 없음"}, status_code=403)
    pid = int(parent_id) if parent_id.strip().isdigit() else None
    set_parent(user_id, pid)
    return JSONResponse({"ok": True})


# ── 초기 관리자 계정 생성 (최초 1회) ─────────────────────
@app.get("/setup", response_class=HTMLResponse)
async def setup_page(request: Request):
    from web.db import get_conn
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if count > 0:
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse(request=request, name="setup.html", context={})


@app.post("/setup")
async def setup(name: str = Form(...), email: str = Form(...), password: str = Form(...)):
    from web.db import get_conn
    with get_conn() as conn:
        if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] > 0:
            return RedirectResponse("/login", status_code=302)
        conn.execute("INSERT INTO users (email, password_hash, name, role) VALUES (?,?,?,?)",
                     (email.strip().lower(), _hash_pw(password), name, "admin"))
    return RedirectResponse("/login", status_code=302)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("web.app:app", host="0.0.0.0", port=8000, reload=True)
